import os
os.environ["APP_ENV"] = "testing"

import unittest
import tempfile
import sys
import io
import openpyxl

os.environ["APP_ENV"] = "testing"
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import modules.notifications.database as notif_db
notif_db.init_notification_database = lambda: None

import services.subscriptions as sub_service
sub_service.assert_subscription_access = lambda *args, **kwargs: None
sub_service.assert_feature_enabled = lambda *args, **kwargs: None
sub_service.lock_and_check_limit = lambda *args, **kwargs: None

from config import Config
Config.DB_TYPE = "sqlite"
Config.APP_ENV = "testing"

from db import init_db, get_conn
from app import app
from modules.reports.routes import upload_csv, download_sample


class TestExcelImportAndFields(unittest.TestCase):
    def setUp(self):
        self.db_fd, self.db_path = tempfile.mkstemp()
        os.environ["APP_ENV"] = "testing"
        os.environ["DATABASE_PATH"] = self.db_path
        import db
        db.DATABASE_PATH = self.db_path
        init_db()

        self.app = app
        self.app.config["TESTING"] = True
        self.app.config["WTF_CSRF_ENABLED"] = False
        self.client = self.app.test_client()

        conn = get_conn()
        cur = conn.cursor()

        # Institute 1
        cur.execute("INSERT OR REPLACE INTO institutes (id, name, short_name, slug, created_at) VALUES (1, 'Inst One', 'IO1', 'inst-one', '2026-01-01 00:00:00')")
        cur.execute("INSERT OR REPLACE INTO branches (id, institute_id, branch_name, branch_code, address, is_active, created_at) VALUES (10, 1, 'Bangalore Central', 'BC01', 'MG Road', 1, '2026-01-01 00:00:00')")
        cur.execute("INSERT OR REPLACE INTO courses (id, institute_id, course_name, duration, fee, course_type, is_active, created_at) VALUES (100, 1, 'Full Stack Web Dev', '90 Days', 15000, 'standard', 1, '2026-01-01 00:00:00')")

        # Institute 2
        cur.execute("INSERT OR REPLACE INTO institutes (id, name, short_name, slug, created_at) VALUES (2, 'Inst Two', 'IO2', 'inst-two', '2026-01-01 00:00:00')")
        cur.execute("INSERT OR REPLACE INTO branches (id, institute_id, branch_name, branch_code, address, is_active, created_at) VALUES (20, 2, 'Mysore North', 'MN01', 'Station Road', 1, '2026-01-01 00:00:00')")

        # Users
        cur.execute("INSERT OR REPLACE INTO users (id, institute_id, username, full_name, password_hash, role, is_active, branch_id, created_at) VALUES (1, 1, 'admin1', 'Admin One', 'hash', 'admin', 1, 10, '2026-01-01 00:00:00')")
        cur.execute("INSERT OR REPLACE INTO users (id, institute_id, username, full_name, password_hash, role, is_active, branch_id, created_at) VALUES (2, 2, 'admin2', 'Admin Two', 'hash', 'admin', 1, 20, '2026-01-01 00:00:00')")

        conn.commit()
        conn.close()

    def tearDown(self):
        os.close(self.db_fd)
        os.unlink(self.db_path)

    def test_excel_sample_download_with_reference_sheets(self):
        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["institute_id"] = 1
            sess["role"] = "admin"
            sess["branch_id"] = 10

        res = self.client.get("/reports/sample/students?file_type=xlsx")
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.mimetype, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        self.assertIn("Students Data", wb.sheetnames)
        self.assertIn("Branches Reference", wb.sheetnames)
        self.assertIn("Courses Reference", wb.sheetnames)

        # Check Branches reference contents
        ws_b = wb["Branches Reference"]
        rows_b = list(ws_b.iter_rows(values_only=True))
        self.assertEqual(rows_b[0], ("Branch ID", "Branch Name", "Branch Code", "Address"))
        self.assertIn((10, "Bangalore Central", "BC01", "MG Road"), rows_b)

    def test_excel_student_import_smart_branch_resolution_and_fields(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students Data"
        headers = [
            "branch_name", "student_code", "full_name", "phone", "email", "gender", "date_of_birth",
            "father_name", "mother_name", "parent_name", "parent_contact",
            "address", "locality", "city", "state", "pincode", "landmark", "alternate_phone", "address_type",
            "education_level", "qualification",
            "tenth_institution", "tenth_board", "tenth_year", "tenth_percentage",
            "puc_institution", "puc_board", "puc_stream", "puc_year", "puc_percentage",
            "degree_institution", "degree_university", "degree_course", "degree_year", "degree_percentage",
            "student_location", "employment_status", "status", "joined_date"
        ]
        ws.append(headers)

        row_data = [
            "Bangalore Central", "", "Rahul Sharma", "9988776655", "rahul@example.com", "Male", "2001-04-10",
            "Ramesh Sharma", "Sunita Sharma", "Ramesh Sharma", "9988776600",
            "123 Brigade Road", "Ashok Nagar", "Bengaluru", "Karnataka", "560025", "Near Brigade Mall", "9988776611", "Home",
            "Undergraduate", "B.Tech",
            "DPS Bengaluru", "CBSE", "2017", "92%",
            "Christ Junior College", "PU Board", "Science", "2019", "90%",
            "BMS College of Engineering", "VTU", "Computer Science", "2023", "8.8 CGPA",
            "urban", "student", "active", "10-01-2026"
        ]
        ws.append(row_data)

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)

        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["institute_id"] = 1
            sess["role"] = "admin"
            sess["branch_id"] = 10

        res = self.client.post(
            "/reports/upload",
            data={
                "table_name": "students",
                "file": (excel_stream, "students.xlsx")
            },
            content_type="multipart/form-data"
        )
        self.assertEqual(res.status_code, 302)

        conn = get_conn()
        student = conn.execute("SELECT * FROM students WHERE phone = '9988776655'").fetchone()
        conn.close()

        self.assertIsNotNone(student)
        self.assertEqual(student["full_name"], "Rahul Sharma")
        self.assertEqual(student["institute_id"], 1)
        self.assertEqual(student["branch_id"], 10) # Resolved from "Bangalore Central"
        self.assertEqual(student["father_name"], "Ramesh Sharma")
        self.assertEqual(student["mother_name"], "Sunita Sharma")
        self.assertEqual(student["pincode"], "560025")
        self.assertEqual(student["degree_university"], "VTU")


if __name__ == "__main__":
    unittest.main()
