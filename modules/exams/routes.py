import random
import json
from datetime import date, datetime

from flask import flash, redirect, render_template, request, session, url_for

from db import get_conn, log_activity
from modules.core.utils import lms_content_manager_required
from . import exams_bp


VALID_CORRECT_OPTIONS = {"A", "B", "C", "D"}


def _get_programs(cur):
    return cur.execute(
        """
            SELECT
                lp.id,
                lp.program_name,
                c.course_name
            FROM lms_programs lp
            LEFT JOIN courses c ON c.id = lp.course_id
            WHERE COALESCE(lp.is_deleted, 0) = 0
            ORDER BY c.course_name COLLATE NOCASE, lp.program_name COLLATE NOCASE
        """
    ).fetchall()


def _get_chapters(cur):
    return cur.execute(
        """
            SELECT
                mc.id,
                mc.title,
                mc.status,
                GROUP_CONCAT(DISTINCT pc.program_id) AS program_ids
            FROM lms_master_chapters mc
            LEFT JOIN lms_program_chapters pc ON pc.master_chapter_id = mc.id
            GROUP BY mc.id
            ORDER BY mc.title COLLATE NOCASE
        """
    ).fetchall()


def _get_questions(cur):
    return cur.execute(
        """
            SELECT
                qb.id,
                qb.chapter_id,
                qb.master_topic_id,
                qb.question_text,
                qb.option_a,
                qb.option_b,
                qb.option_c,
                qb.option_d,
                qb.correct_option,
                qb.question_type,
                COALESCE(mc.title, 'Unknown Chapter') AS chapter_name,
                mt.title AS topic_name,
                GROUP_CONCAT(DISTINCT pc.program_id) AS program_ids
            FROM lms_question_bank qb
            LEFT JOIN lms_master_chapters mc ON mc.id = qb.chapter_id
            LEFT JOIN lms_master_topics mt ON mt.id = qb.master_topic_id
            LEFT JOIN lms_program_chapters pc ON pc.master_chapter_id = mc.id
            GROUP BY qb.id
            ORDER BY qb.id DESC
        """
    ).fetchall()


def _get_question_bank_context(cur):
    return _get_programs(cur), _get_chapters(cur), _get_questions(cur)


def _question_form_fields():
    return {
        "chapter_id": request.form.get("chapter_id", "").strip(),
        "question_text": request.form.get("question_text", "").strip(),
        "option_a": request.form.get("option_a", "").strip(),
        "option_b": request.form.get("option_b", "").strip(),
        "option_c": request.form.get("option_c", "").strip(),
        "option_d": request.form.get("option_d", "").strip(),
        "correct_option": request.form.get("correct_option", "").strip().upper(),
        "master_topic_id": request.form.get("master_topic_id", "").strip(),
    }


def _validate_question_fields(fields):
    # Exclude master_topic_id from the required fields check since it is optional
    required_fields = {k: v for k, v in fields.items() if k != "master_topic_id"}
    if any(not value for value in required_fields.values()):
        return "All question fields are required."

    if not fields["chapter_id"].isdigit():
        return "Please select a valid chapter."

    if fields["master_topic_id"] and not fields["master_topic_id"].isdigit():
        return "Please select a valid topic."

    if fields["correct_option"] not in VALID_CORRECT_OPTIONS:
        return "Correct answer must be A, B, C, or D."

    return None


def _student_required_redirect():
    if "student_id" not in session:
        flash("Please login first.", "warning")
        return redirect(url_for("students.login"))
    return None


def _get_student_programs(cur, student_id):
    if session.get("demo_mode"):
        return cur.execute(
            """
                SELECT lp.id, lp.program_name, c.course_name
                FROM lms_programs lp
                LEFT JOIN courses c ON c.id = lp.course_id
                WHERE lp.is_active = 1
                  AND COALESCE(lp.is_deleted, 0) = 0
                ORDER BY COALESCE(c.course_name, lp.program_name) COLLATE NOCASE,
                         lp.program_name COLLATE NOCASE
            """
        ).fetchall()

    return cur.execute(
        """
            SELECT DISTINCT lp.id, lp.program_name, c.course_name
            FROM lms_programs lp
            LEFT JOIN courses c ON c.id = lp.course_id
            WHERE lp.is_active = 1
              AND COALESCE(lp.is_deleted, 0) = 0
              AND (
                EXISTS (
                    SELECT 1
                    FROM lms_student_program_access spa
                    WHERE spa.student_id = ?
                      AND spa.program_id = lp.id
                      AND spa.is_active = 1
                      AND COALESCE(spa.access_status, 'active') = 'active'
                      AND (spa.access_end_date IS NULL OR spa.access_end_date >= date('now'))
                )
                OR EXISTS (
                    SELECT 1
                    FROM lms_batch_program_access bpa
                    JOIN student_batches sb ON bpa.batch_id = sb.batch_id
                    WHERE sb.student_id = ?
                      AND bpa.program_id = lp.id
                      AND bpa.is_active = 1
                      AND (bpa.access_end_date IS NULL OR bpa.access_end_date >= date('now'))
                )
                OR EXISTS (
                    SELECT 1
                    FROM invoices i
                    JOIN invoice_items ii ON ii.invoice_id = i.id
                    WHERE i.student_id = ?
                      AND ii.course_id = lp.course_id
                      AND lp.course_id IS NOT NULL
                )
                OR EXISTS (
                    SELECT 1
                    FROM invoices i
                    JOIN invoice_items ii ON ii.invoice_id = i.id
                    JOIN lms_course_program_map cpm
                         ON cpm.course_id = ii.course_id AND cpm.program_id = lp.id
                    WHERE i.student_id = ?
                )
              )
            ORDER BY COALESCE(c.course_name, lp.program_name) COLLATE NOCASE,
                     lp.program_name COLLATE NOCASE
        """,
        (student_id, student_id, student_id, student_id),
    ).fetchall()


def _student_has_program_access(cur, student_id, program_id):
    if session.get("demo_mode"):
        return True

    return cur.execute(
        """
            SELECT 1
            FROM lms_programs lp
            WHERE lp.id = ?
              AND lp.is_active = 1
              AND COALESCE(lp.is_deleted, 0) = 0
              AND (
                EXISTS (
                    SELECT 1
                    FROM lms_student_program_access spa
                    WHERE spa.student_id = ?
                      AND spa.program_id = lp.id
                      AND spa.is_active = 1
                      AND COALESCE(spa.access_status, 'active') = 'active'
                      AND (spa.access_end_date IS NULL OR spa.access_end_date >= date('now'))
                )
                OR EXISTS (
                    SELECT 1
                    FROM lms_batch_program_access bpa
                    JOIN student_batches sb ON bpa.batch_id = sb.batch_id
                    WHERE sb.student_id = ?
                      AND bpa.program_id = lp.id
                      AND bpa.is_active = 1
                      AND (bpa.access_end_date IS NULL OR bpa.access_end_date >= date('now'))
                )
                OR EXISTS (
                    SELECT 1
                    FROM invoices i
                    JOIN invoice_items ii ON ii.invoice_id = i.id
                    WHERE i.student_id = ?
                      AND ii.course_id = lp.course_id
                      AND lp.course_id IS NOT NULL
                )
                OR EXISTS (
                    SELECT 1
                    FROM invoices i
                    JOIN invoice_items ii ON ii.invoice_id = i.id
                    JOIN lms_course_program_map cpm
                         ON cpm.course_id = ii.course_id AND cpm.program_id = lp.id
                    WHERE i.student_id = ?
                )
              )
        """,
        (program_id, student_id, student_id, student_id, student_id),
    ).fetchone() is not None


def _table_columns(cur, table_name):
    return {row["name"] for row in cur.execute(f"PRAGMA table_info({table_name})").fetchall()}


def _final_exam_syllabus_check(cur, student_id, program_id):
    master_total = cur.execute(
        """
            SELECT COUNT(DISTINCT mt.id) AS total
            FROM lms_program_chapters pc
            JOIN lms_master_chapters mc ON mc.id = pc.master_chapter_id
            JOIN lms_master_topics mt ON mt.master_chapter_id = mc.id
            WHERE pc.program_id = ?
              AND COALESCE(pc.is_visible, 1) = 1
              AND COALESCE(mc.status, 'active') = 'active'
              AND COALESCE(mt.status, 'active') = 'active'
        """,
        (program_id,),
    ).fetchone()["total"] or 0

    if master_total:
        completed = cur.execute(
            """
                SELECT COUNT(DISTINCT mt.id) AS completed
                FROM lms_program_chapters pc
                JOIN lms_master_chapters mc ON mc.id = pc.master_chapter_id
                JOIN lms_master_topics mt ON mt.master_chapter_id = mc.id
                JOIN lms_master_topic_progress mp
                  ON mp.master_topic_id = mt.id
                 AND mp.program_id = pc.program_id
                 AND mp.student_id = ?
                 AND mp.is_completed = 1
                WHERE pc.program_id = ?
                  AND COALESCE(pc.is_visible, 1) = 1
                  AND COALESCE(mc.status, 'active') = 'active'
                  AND COALESCE(mt.status, 'active') = 'active'
            """,
            (student_id, program_id),
        ).fetchone()["completed"] or 0
        return {
            "total": master_total,
            "completed": completed,
            "passed": completed >= master_total,
        }

    legacy_total = cur.execute(
        """
            SELECT COUNT(DISTINCT lt.id) AS total
            FROM lms_chapters lc
            JOIN lms_topics lt ON lt.chapter_id = lc.id
            WHERE lc.program_id = ?
              AND COALESCE(lc.is_active, 1) = 1
              AND COALESCE(lt.is_active, 1) = 1
        """,
        (program_id,),
    ).fetchone()["total"] or 0
    completed = cur.execute(
        """
            SELECT COUNT(DISTINCT lt.id) AS completed
            FROM lms_chapters lc
            JOIN lms_topics lt ON lt.chapter_id = lc.id
            JOIN lms_topic_progress tp
              ON tp.topic_id = lt.id
             AND tp.student_id = ?
             AND tp.is_completed = 1
            WHERE lc.program_id = ?
              AND COALESCE(lc.is_active, 1) = 1
              AND COALESCE(lt.is_active, 1) = 1
        """,
        (student_id, program_id),
    ).fetchone()["completed"] or 0
    return {
        "total": legacy_total,
        "completed": completed,
        "passed": legacy_total > 0 and completed >= legacy_total,
    }


def _final_exam_assignment_check(cur, student_id, program_id):
    submission_columns = _table_columns(cur, "lms_assignment_submissions")
    review_column = "s.review_status" if "review_status" in submission_columns else "s.status"
    latest_clause = "AND COALESCE(s.is_latest, 1) = 1" if "is_latest" in submission_columns else ""

    total = cur.execute(
        """
            SELECT COUNT(DISTINCT a.id) AS total
            FROM lms_assignments a
            JOIN lms_master_topics mt ON mt.id = a.master_topic_id
            JOIN lms_master_chapters mc ON mc.id = mt.master_chapter_id
            JOIN lms_program_chapters pc ON pc.master_chapter_id = mc.id
            WHERE pc.program_id = ?
              AND COALESCE(pc.is_visible, 1) = 1
              AND COALESCE(mc.status, 'active') = 'active'
              AND COALESCE(mt.status, 'active') = 'active'
        """,
        (program_id,),
    ).fetchone()["total"] or 0

    submitted = cur.execute(
        f"""
            SELECT COUNT(DISTINCT a.id) AS submitted
            FROM lms_assignments a
            JOIN lms_master_topics mt ON mt.id = a.master_topic_id
            JOIN lms_master_chapters mc ON mc.id = mt.master_chapter_id
            JOIN lms_program_chapters pc ON pc.master_chapter_id = mc.id
            JOIN lms_assignment_submissions s
              ON s.assignment_id = a.id
             AND s.student_id = ?
             {latest_clause}
             AND COALESCE({review_column}, s.status, 'submitted') IN ('submitted', 'accepted', 'reviewed')
            WHERE pc.program_id = ?
              AND COALESCE(pc.is_visible, 1) = 1
              AND COALESCE(mc.status, 'active') = 'active'
              AND COALESCE(mt.status, 'active') = 'active'
        """,
        (student_id, program_id),
    ).fetchone()["submitted"] or 0

    return {
        "total": total,
        "submitted": submitted,
        "missing": max(total - submitted, 0),
        "passed": submitted >= total,
    }


def _final_exam_dues_check(cur, student_id):
    balance = cur.execute(
        """
            SELECT COALESCE(SUM(
                CASE
                    WHEN i.total_amount 
                         - COALESCE((SELECT SUM(r.amount_received) FROM receipts r WHERE r.invoice_id = i.id), 0)
                         - COALESCE((SELECT SUM(w.amount_written_off) FROM bad_debt_writeoffs w WHERE w.invoice_id = i.id), 0) > 0
                    THEN i.total_amount 
                         - COALESCE((SELECT SUM(r.amount_received) FROM receipts r WHERE r.invoice_id = i.id), 0)
                         - COALESCE((SELECT SUM(w.amount_written_off) FROM bad_debt_writeoffs w WHERE w.invoice_id = i.id), 0)
                    ELSE 0
                END
            ), 0) AS balance
            FROM invoices i
            WHERE i.student_id = ?
              AND COALESCE(i.status, '') NOT IN ('cancelled', 'write_off')
        """,
        (student_id,),
    ).fetchone()["balance"] or 0
    balance = round(float(balance), 2)
    return {
        "balance": balance,
        "passed": balance <= 0,
    }


def _final_exam_checks(cur, student_id, program_id):
    syllabus = _final_exam_syllabus_check(cur, student_id, program_id)
    assignments = _final_exam_assignment_check(cur, student_id, program_id)
    dues = _final_exam_dues_check(cur, student_id)
    return {
        "syllabus": syllabus,
        "assignments": assignments,
        "dues": dues,
        "all_passed": syllabus["passed"] and assignments["passed"] and dues["passed"],
    }


def _latest_final_exam_application(cur, student_id, program_id):
    return cur.execute(
        """
            SELECT *
            FROM lms_final_exam_applications
            WHERE student_id = ?
              AND course_id = ?
            ORDER BY applied_on DESC, id DESC
            LIMIT 1
        """,
        (student_id, program_id),
    ).fetchone()


def _get_final_exam_application(cur, application_id):
    return cur.execute(
        """
            SELECT
                app.*,
                s.student_code,
                s.full_name AS current_student_name,
                s.phone AS current_student_phone,
                s.date_of_birth AS current_student_dob,
                lp.program_name,
                c.course_name
            FROM lms_final_exam_applications app
            JOIN students s ON s.id = app.student_id
            LEFT JOIN lms_programs lp ON lp.id = app.course_id
            LEFT JOIN courses c ON c.id = lp.course_id
            WHERE app.id = ?
        """,
        (application_id,),
    ).fetchone()


def _get_student_final_exam_application(cur, application_id, student_id):
    return cur.execute(
        """
            SELECT
                app.*,
                lp.program_name,
                c.course_name
            FROM lms_final_exam_applications app
            LEFT JOIN lms_programs lp ON lp.id = app.course_id
            LEFT JOIN courses c ON c.id = lp.course_id
            WHERE app.id = ?
              AND app.student_id = ?
        """,
        (application_id, student_id),
    ).fetchone()


def _get_final_exam_attempt(cur, application_id, student_id):
    return cur.execute(
        """
            SELECT *
            FROM lms_final_exam_attempts
            WHERE application_id = ?
              AND student_id = ?
        """,
        (application_id, student_id),
    ).fetchone()


def _is_final_exam_unlocked(application):
    if not application or application["status"] != "APPROVED":
        return False
    try:
        return date.fromisoformat(application["requested_exam_date"]) <= date.today()
    except (TypeError, ValueError):
        return False


def _get_final_exam_questions(cur, program_id):
    return cur.execute(
        """
            SELECT DISTINCT
                qb.id,
                qb.chapter_id,
                qb.question_text,
                qb.option_a,
                qb.option_b,
                qb.option_c,
                qb.option_d,
                qb.question_type
            FROM lms_question_bank qb
            JOIN lms_program_chapters pc ON pc.master_chapter_id = qb.chapter_id
            JOIN lms_master_chapters mc ON mc.id = qb.chapter_id
            WHERE pc.program_id = ?
              AND COALESCE(pc.is_visible, 1) = 1
              AND COALESCE(mc.status, 'active') = 'active'
            ORDER BY qb.id
        """,
        (program_id,),
    ).fetchall()


@exams_bp.route("/lms_admin/questions/manage", methods=["GET"])
@lms_content_manager_required
def manage_questions():
    conn = get_conn()
    try:
        cur = conn.cursor()
        programs = _get_programs(cur)
        chapters = _get_chapters(cur)
        topics = cur.execute(
            "SELECT id, master_chapter_id, title FROM lms_master_topics WHERE status = 'active' ORDER BY topic_order"
        ).fetchall()
        return render_template(
            "exams/admin_manage_questions.html",
            programs=programs,
            chapters=chapters,
            topics=topics,
            question=None,
            form_action=url_for("exams.add_question"),
            submit_label="Add Question",
            page_heading="Add Question",
            page_subtitle="Create MCQ questions linked to reusable LMS chapters.",
        )
    finally:
        conn.close()


@exams_bp.route("/lms_admin/questions/view", methods=["GET"])
@lms_content_manager_required
def view_questions():
    conn = get_conn()
    try:
        cur = conn.cursor()
        programs, chapters, questions = _get_question_bank_context(cur)
        return render_template(
            "exams/admin_view_questions.html",
            programs=programs,
            chapters=chapters,
            questions=questions,
        )
    finally:
        conn.close()


@exams_bp.route("/lms_admin/questions/edit/<int:question_id>", methods=["GET"])
@lms_content_manager_required
def edit_question(question_id):
    conn = get_conn()
    try:
        cur = conn.cursor()
        question = cur.execute(
            "SELECT * FROM lms_question_bank WHERE id = ?",
            (question_id,),
        ).fetchone()
        if not question:
            flash("Question not found.", "warning")
            return redirect(url_for("exams.view_questions"))

        programs = _get_programs(cur)
        chapters = _get_chapters(cur)
        topics = cur.execute(
            "SELECT id, master_chapter_id, title FROM lms_master_topics WHERE status = 'active' ORDER BY topic_order"
        ).fetchall()
        return render_template(
            "exams/admin_manage_questions.html",
            programs=programs,
            chapters=chapters,
            topics=topics,
            question=question,
            form_action=url_for("exams.update_question", question_id=question_id),
            submit_label="Update Question",
            page_heading=f"Edit Question #{question_id}",
            page_subtitle="Update the question, options, chapter, and correct answer.",
        )
    finally:
        conn.close()


@exams_bp.route("/lms_admin/questions/add", methods=["POST"])
@lms_content_manager_required
def add_question():
    fields = _question_form_fields()
    error = _validate_question_fields(fields)
    if error:
        flash(error, "danger")
        return redirect(url_for("exams.manage_questions"))

    chapter_id = int(fields["chapter_id"])
    master_topic_id = int(fields["master_topic_id"]) if fields["master_topic_id"] else None
    conn = get_conn()
    try:
        cur = conn.cursor()
        chapter = cur.execute(
            "SELECT id, title FROM lms_master_chapters WHERE id = ?",
            (chapter_id,),
        ).fetchone()
        if not chapter:
            flash("Selected chapter was not found.", "danger")
            return redirect(url_for("exams.manage_questions"))

        cur.execute(
            """
                INSERT INTO lms_question_bank (
                    chapter_id,
                    master_topic_id,
                    question_text,
                    option_a,
                    option_b,
                    option_c,
                    option_d,
                    correct_option,
                    question_type
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'MCQ')
            """,
            (
                chapter_id,
                master_topic_id,
                fields["question_text"],
                fields["option_a"],
                fields["option_b"],
                fields["option_c"],
                fields["option_d"],
                fields["correct_option"],
            ),
        )
        question_id = cur.lastrowid

        log_activity(
            user_id=session.get("user_id"),
            branch_id=session.get("branch_id"),
            action_type="create",
            module_name="lms_question_bank",
            record_id=question_id,
            description=f"Created question for chapter: {chapter['title']}",
            conn=conn,
        )
        conn.commit()
        flash("Question added successfully.", "success")
    finally:
        conn.close()

    return redirect(url_for("exams.view_questions"))


@exams_bp.route("/lms_admin/questions/update/<int:question_id>", methods=["POST"])
@lms_content_manager_required
def update_question(question_id):
    fields = _question_form_fields()
    error = _validate_question_fields(fields)
    if error:
        flash(error, "danger")
        return redirect(url_for("exams.edit_question", question_id=question_id))

    chapter_id = int(fields["chapter_id"])
    master_topic_id = int(fields["master_topic_id"]) if fields["master_topic_id"] else None
    conn = get_conn()
    try:
        cur = conn.cursor()
        question = cur.execute(
            "SELECT id FROM lms_question_bank WHERE id = ?",
            (question_id,),
        ).fetchone()
        if not question:
            flash("Question not found.", "warning")
            return redirect(url_for("exams.view_questions"))

        chapter = cur.execute(
            "SELECT id, title FROM lms_master_chapters WHERE id = ?",
            (chapter_id,),
        ).fetchone()
        if not chapter:
            flash("Selected chapter was not found.", "danger")
            return redirect(url_for("exams.edit_question", question_id=question_id))

        cur.execute(
            """
                UPDATE lms_question_bank
                SET
                    chapter_id = ?,
                    master_topic_id = ?,
                    question_text = ?,
                    option_a = ?,
                    option_b = ?,
                    option_c = ?,
                    option_d = ?,
                    correct_option = ?,
                    question_type = 'MCQ'
                WHERE id = ?
            """,
            (
                chapter_id,
                master_topic_id,
                fields["question_text"],
                fields["option_a"],
                fields["option_b"],
                fields["option_c"],
                fields["option_d"],
                fields["correct_option"],
                question_id,
            ),
        )

        log_activity(
            user_id=session.get("user_id"),
            branch_id=session.get("branch_id"),
            action_type="update",
            module_name="lms_question_bank",
            record_id=question_id,
            description=f"Updated question for chapter: {chapter['title']}",
            conn=conn,
        )
        conn.commit()
        flash("Question updated successfully.", "success")
    finally:
        conn.close()

    return redirect(url_for("exams.view_questions"))


@exams_bp.route("/lms_admin/questions/delete/<int:question_id>", methods=["GET", "POST"])
@lms_content_manager_required
def delete_question(question_id):
    conn = get_conn()
    try:
        cur = conn.cursor()
        question = cur.execute(
            "SELECT id, question_text FROM lms_question_bank WHERE id = ?",
            (question_id,),
        ).fetchone()
        if not question:
            flash("Question not found.", "warning")
            return redirect(url_for("exams.view_questions"))

        cur.execute("DELETE FROM lms_question_bank WHERE id = ?", (question_id,))
        log_activity(
            user_id=session.get("user_id"),
            branch_id=session.get("branch_id"),
            action_type="delete",
            module_name="lms_question_bank",
            record_id=question_id,
            description=f"Deleted question: {question['question_text'][:80]}",
            conn=conn,
        )
        conn.commit()
        flash("Question deleted successfully.", "success")
    finally:
        conn.close()

    return redirect(url_for("exams.view_questions"))


@exams_bp.route("/lms_admin/questions/import", methods=["POST"])
@lms_content_manager_required
def import_questions():
    """Import multiple questions from a CSV or Excel file into a selected chapter."""
    import csv
    import io
    from openpyxl import load_workbook
    
    chapter_id = request.form.get("chapter_id", type=int)
    if not chapter_id:
        flash("Please select a target chapter.", "danger")
        return redirect(url_for("exams.view_questions"))
        
    uploaded_file = request.files.get("csv_file") # keeping input name "csv_file" for form compatibility
    if not uploaded_file or not uploaded_file.filename:
        flash("Please upload a file.", "danger")
        return redirect(url_for("exams.view_questions"))
        
    filename = uploaded_file.filename.lower()
    is_xlsx = filename.endswith('.xlsx')
    is_csv = filename.endswith('.csv')
    
    if not is_xlsx and not is_csv:
        flash("Please upload a valid CSV (.csv) or Excel (.xlsx) file.", "danger")
        return redirect(url_for("exams.view_questions"))
        
    conn = get_conn()
    try:
        cur = conn.cursor()
        # Verify chapter exists
        chapter = cur.execute("SELECT id, title FROM lms_master_chapters WHERE id = ?", (chapter_id,)).fetchone()
        if not chapter:
            flash("Selected chapter not found.", "danger")
            return redirect(url_for("exams.view_questions"))
            
        # Fetch active topics for lookup mapping
        topics = cur.execute(
            "SELECT id, title FROM lms_master_topics WHERE master_chapter_id = ? AND status = 'active'",
            (chapter_id,)
        ).fetchall()
        topic_lookup = {t["title"].strip().lower(): t["id"] for t in topics}
        
        rows_to_import = []
        
        if is_xlsx:
            # Parse Excel using openpyxl
            wb = load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            
            iter_rows = ws.iter_rows(values_only=True)
            headers = next(iter_rows, None)
            if not headers:
                flash("The Excel sheet is empty.", "danger")
                return redirect(url_for("exams.view_questions"))
                
            headers = [str(h).strip().lower() for h in headers if h is not None]
            
            col_map = {
                "topic_title": ["topic title", "topic", "topic_title"],
                "question_text": ["question_text", "question", "question text"],
                "option_a": ["option_a", "option a", "a"],
                "option_b": ["option_b", "option b", "b"],
                "option_c": ["option_c", "option c", "c"],
                "option_d": ["option_d", "option d", "d"],
                "correct_option": ["correct_option", "correct option", "correct answer", "correct"]
            }
            
            indices = {}
            for col, aliases in col_map.items():
                idx = -1
                for alias in aliases:
                    if alias in headers:
                        idx = headers.index(alias)
                        break
                if col != "topic_title" and idx == -1:
                    flash(f"Required column '{col}' was not found in the Excel sheet. Headers must contain: Question Text, Option A, Option B, Option C, Option D, Correct Answer.", "danger")
                    return redirect(url_for("exams.view_questions"))
                indices[col] = idx
                
            row_num = 1
            for row in iter_rows:
                row_num += 1
                if not row or not any(x is not None and str(x).strip() for x in row):
                    continue
                    
                def get_val(col_name):
                    idx = indices.get(col_name, -1)
                    if idx != -1 and idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                    return ""
                    
                q_text = get_val("question_text")
                opt_a = get_val("option_a")
                opt_b = get_val("option_b")
                opt_c = get_val("option_c")
                opt_d = get_val("option_d")
                corr = get_val("correct_option").upper()
                t_title = get_val("topic_title")
                
                if not q_text or not opt_a or not opt_b or not opt_c or not opt_d or not corr:
                    flash(f"Error in row {row_num}: Question text, options (A-D), and correct answer must all be filled.", "danger")
                    return redirect(url_for("exams.view_questions"))
                    
                if corr not in ('A', 'B', 'C', 'D'):
                    flash(f"Error in row {row_num}: Correct answer must be A, B, C, or D. Got '{corr}'.", "danger")
                    return redirect(url_for("exams.view_questions"))
                    
                # Resolve topic_id
                master_topic_id = None
                if t_title:
                    master_topic_id = topic_lookup.get(t_title.lower())
                    if not master_topic_id:
                        flash(f"Error in row {row_num}: Topic '{t_title}' not found in the selected chapter. Please check or add the topic first.", "danger")
                        return redirect(url_for("exams.view_questions"))
                        
                rows_to_import.append((master_topic_id, q_text, opt_a, opt_b, opt_c, opt_d, corr))
                
        else:
            # Parse CSV
            stream = io.StringIO(uploaded_file.stream.read().decode("utf-8-sig"), newline=None)
            reader = csv.reader(stream)
            
            headers = next(reader, None)
            if not headers:
                flash("The CSV file is empty.", "danger")
                return redirect(url_for("exams.view_questions"))
                
            headers = [h.strip().lower() for h in headers]
            
            col_map = {
                "topic_title": ["topic title", "topic", "topic_title"],
                "question_text": ["question_text", "question", "question text"],
                "option_a": ["option_a", "option a", "a"],
                "option_b": ["option_b", "option b", "b"],
                "option_c": ["option_c", "option c", "c"],
                "option_d": ["option_d", "option d", "d"],
                "correct_option": ["correct_option", "correct option", "correct answer", "correct"]
            }
            
            indices = {}
            for col, aliases in col_map.items():
                idx = -1
                for alias in aliases:
                    if alias in headers:
                        idx = headers.index(alias)
                        break
                if col != "topic_title" and idx == -1:
                    flash(f"Required column '{col}' was not found in the CSV. Headers must contain: Question Text, Option A, Option B, Option C, Option D, Correct Answer.", "danger")
                    return redirect(url_for("exams.view_questions"))
                indices[col] = idx
                
            row_num = 1
            for row in reader:
                row_num += 1
                if not row or not any(field.strip() for field in row):
                    continue
                    
                def get_val(col_name):
                    idx = indices.get(col_name, -1)
                    if idx != -1 and idx < len(row):
                        return row[idx].strip()
                    return ""
                    
                q_text = get_val("question_text")
                opt_a = get_val("option_a")
                opt_b = get_val("option_b")
                opt_c = get_val("option_c")
                opt_d = get_val("option_d")
                corr = get_val("correct_option").upper()
                t_title = get_val("topic_title")
                
                if not q_text or not opt_a or not opt_b or not opt_c or not opt_d or not corr:
                    flash(f"Error in row {row_num}: Question text, options (A-D), and correct answer must all be filled.", "danger")
                    return redirect(url_for("exams.view_questions"))
                    
                if corr not in ('A', 'B', 'C', 'D'):
                    flash(f"Error in row {row_num}: Correct answer must be A, B, C, or D. Got '{corr}'.", "danger")
                    return redirect(url_for("exams.view_questions"))
                    
                # Resolve topic_id
                master_topic_id = None
                if t_title:
                    master_topic_id = topic_lookup.get(t_title.lower())
                    if not master_topic_id:
                        flash(f"Error in row {row_num}: Topic '{t_title}' not found in the selected chapter. Please check or add the topic first.", "danger")
                        return redirect(url_for("exams.view_questions"))
                        
                rows_to_import.append((master_topic_id, q_text, opt_a, opt_b, opt_c, opt_d, corr))
                
        # Perform insertion of collected rows
        imported_count = 0
        for r in rows_to_import:
            cur.execute("""
                INSERT INTO lms_question_bank (
                    chapter_id, master_topic_id, question_text, option_a, option_b, option_c, option_d, correct_option, question_type
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'MCQ')
            """, (chapter_id, r[0], r[1], r[2], r[3], r[4], r[5], r[6]))
            imported_count += 1
            
        conn.commit()
        log_activity(
            user_id=session.get("user_id"),
            branch_id=session.get("branch_id"),
            action_type="create",
            module_name="lms_question_bank",
            record_id=None,
            description=f"Imported {imported_count} questions from file into chapter: {chapter['title']}",
            conn=conn,
        )
        flash(f"Successfully imported {imported_count} questions into '{chapter['title']}'.", "success")
        
    except Exception as e:
        flash(f"Error importing file: {str(e)}", "danger")
    finally:
        conn.close()
        
    return redirect(url_for("exams.view_questions"))


@exams_bp.route("/lms_admin/questions/download-template", methods=["GET"])
@lms_content_manager_required
def download_template():
    """Download a dynamic questions Excel template (.xlsx) for a selected chapter."""
    import io
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from flask import Response

    chapter_id = request.args.get("chapter_id", type=int)
    if not chapter_id:
        return "Chapter ID is required to generate the template.", 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        chapter = cur.execute(
            "SELECT id, title FROM lms_master_chapters WHERE id = ?",
            (chapter_id,),
        ).fetchone()
        if not chapter:
            return "Chapter not found.", 404

        # Fetch active topics for this chapter
        topics = cur.execute(
            """
                SELECT id, title
                FROM lms_master_topics
                WHERE master_chapter_id = ? AND status = 'active'
                ORDER BY topic_order
            """,
            (chapter_id,),
        ).fetchall()
    finally:
        conn.close()

    wb = Workbook()
    
    # 1. Sheet 1: Questions
    ws_q = wb.active
    ws_q.title = "Questions"
    headers = [
        "Topic Title",
        "Question Text",
        "Option A",
        "Option B",
        "Option C",
        "Option D",
        "Correct Answer",
    ]
    ws_q.append(headers)

    # 2. Sheet 2: Valid Topics
    ws_t = wb.create_sheet(title="Valid Topics")
    topic_titles = [t["title"] for t in topics]
    for title in topic_titles:
        ws_t.append([title])

    # Sample rows on Sheet 1 (if topics exist)
    sample_topic = topic_titles[0] if topic_titles else "General (No Topic)"
    ws_q.append([
        sample_topic,
        "What is the capital of France?",
        "Berlin",
        "Madrid",
        "Paris",
        "Rome",
        "C"
    ])
    ws_q.append([
        sample_topic,
        "Which keyword is used to define a function in Python?",
        "func",
        "def",
        "function",
        "define",
        "B"
    ])

    # 3. Data Validation Dropdowns
    if topic_titles:
        # Reference the Valid Topics sheet range
        formula_topics = f"'Valid Topics'!$A$1:$A${len(topic_titles)}"
        dv_topic = DataValidation(
            type="list",
            formula1=formula_topics,
            allow_blank=True
        )
        dv_topic.error = "Please select a valid topic from the list."
        dv_topic.errorTitle = "Invalid Topic"
        dv_topic.prompt = "Select the topic for this question"
        dv_topic.promptTitle = "Topic Title"
        ws_q.add_data_validation(dv_topic)
        # Apply validation from A2 down to A500
        dv_topic.add("A2:A500")

    # Excel Data validation for Correct Answer (Column G)
    dv_correct = DataValidation(
        type="list",
        formula1='"A,B,C,D"',
        allow_blank=False
    )
    dv_correct.error = "Correct Answer must be A, B, C, or D."
    dv_correct.errorTitle = "Invalid Option"
    dv_correct.prompt = "Choose the correct answer option"
    dv_correct.promptTitle = "Correct Answer"
    ws_q.add_data_validation(dv_correct)
    # Apply validation from G2 down to G500
    dv_correct.add("G2:G500")

    # Save to BytesIO
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"template_chapter_{chapter_id}.xlsx"
    resp = Response(
        out.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


@exams_bp.route("/lms_admin/questions/sample-csv", methods=["GET"])
@lms_content_manager_required
def download_sample_csv():
    """Download a sample questions CSV template."""
    import io
    import csv
    from flask import Response
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Question Text", "Option A", "Option B", "Option C", "Option D", "Correct Answer"])
    writer.writerow(["What is the capital of France?", "Berlin", "Madrid", "Paris", "Rome", "C"])
    writer.writerow(["Which keyword is used to define a function in Python?", "func", "def", "function", "define", "B"])
    
    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=sample_questions.csv"
    return response


@exams_bp.route("/student/final-exam/apply", methods=["GET", "POST"])
def final_exam_apply():
    auth_redirect = _student_required_redirect()
    if auth_redirect:
        return auth_redirect

    student_id = session["student_id"]
    conn = get_conn()
    try:
        cur = conn.cursor()
        student = cur.execute(
            """
                SELECT id, student_code, full_name, phone, date_of_birth
                FROM students
                WHERE id = ?
            """,
            (student_id,),
        ).fetchone()
        programs = _get_student_programs(cur, student_id)

        if request.method == "POST":
            program_id = request.form.get("program_id", type=int)
            if not program_id or not _student_has_program_access(cur, student_id, program_id):
                flash("Please select a valid LMS program.", "danger")
                return redirect(url_for("exams.final_exam_apply"))

            checks = _final_exam_checks(cur, student_id, program_id)
            if not checks["all_passed"]:
                flash("Final exam application is blocked until all requirements are complete.", "warning")
                return redirect(url_for("exams.final_exam_apply", program_id=program_id))

            existing = _latest_final_exam_application(cur, student_id, program_id)
            if existing and existing["status"] in ("PENDING", "APPROVED"):
                flash("You already have a final exam application for this program.", "warning")
                return redirect(url_for("exams.final_exam_apply", program_id=program_id))

            verified_name = request.form.get("verified_name", "").strip()
            verified_phone = request.form.get("verified_phone", "").strip()
            verified_dob = request.form.get("verified_dob", "").strip()
            requested_exam_date = request.form.get("requested_exam_date", "").strip()

            if not all([verified_name, verified_phone, verified_dob, requested_exam_date]):
                flash("Please verify your name, phone, date of birth, and exam date.", "danger")
                return redirect(url_for("exams.final_exam_apply", program_id=program_id))

            try:
                requested_date = date.fromisoformat(requested_exam_date)
            except ValueError:
                flash("Please choose a valid exam date.", "danger")
                return redirect(url_for("exams.final_exam_apply", program_id=program_id))

            if requested_date < date.today():
                flash("Requested exam date cannot be in the past.", "danger")
                return redirect(url_for("exams.final_exam_apply", program_id=program_id))

            cur.execute(
                """
                    INSERT INTO lms_final_exam_applications (
                        student_id,
                        course_id,
                        verified_name,
                        verified_phone,
                        verified_dob,
                        requested_exam_date,
                        status,
                        applied_on
                    ) VALUES (?, ?, ?, ?, ?, ?, 'PENDING', ?)
                """,
                (
                    student_id,
                    program_id,
                    verified_name,
                    verified_phone,
                    verified_dob,
                    requested_exam_date,
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
            conn.commit()
            flash("Final exam application submitted for staff review.", "success")
            return redirect(url_for("exams.final_exam_apply", program_id=program_id))

        selected_program_id = request.args.get("program_id", type=int)
        if not selected_program_id and len(programs) == 1:
            selected_program_id = programs[0]["id"]
        if selected_program_id and not any(program["id"] == selected_program_id for program in programs):
            selected_program_id = None

        selected_program = None
        checks = None
        application = None
        if selected_program_id:
            selected_program = next(
                (program for program in programs if program["id"] == selected_program_id),
                None,
            )
            checks = _final_exam_checks(cur, student_id, selected_program_id)
            application = _latest_final_exam_application(cur, student_id, selected_program_id)

        return render_template(
            "exams/student_final_exam_apply.html",
            student=student,
            programs=programs,
            selected_program=selected_program,
            selected_program_id=selected_program_id,
            checks=checks,
            application=application,
            today=date.today().isoformat(),
        )
    finally:
        conn.close()


@exams_bp.route("/lms_admin/final-exam/applications", methods=["GET"])
@lms_content_manager_required
def final_exam_applications():
    conn = get_conn()
    try:
        applications = conn.execute(
            """
                SELECT
                    app.*,
                    s.student_code,
                    s.full_name AS current_student_name,
                    s.phone AS current_student_phone,
                    lp.program_name,
                    c.course_name
                FROM lms_final_exam_applications app
                JOIN students s ON s.id = app.student_id
                LEFT JOIN lms_programs lp ON lp.id = app.course_id
                LEFT JOIN courses c ON c.id = lp.course_id
                ORDER BY
                    CASE app.status
                        WHEN 'PENDING' THEN 0
                        WHEN 'APPROVED' THEN 1
                        ELSE 2
                    END,
                    app.applied_on DESC
            """
        ).fetchall()
    finally:
        conn.close()

    return render_template(
        "exams/admin_final_exam_applications.html",
        applications=applications,
        today=date.today().isoformat(),
    )


@exams_bp.route("/lms_admin/final-exam/applications/<int:application_id>/approve", methods=["POST"])
@lms_content_manager_required
def approve_final_exam_application(application_id):
    conn = get_conn()
    try:
        cur = conn.cursor()
        application = _get_final_exam_application(cur, application_id)
        if not application:
            flash("Final exam application not found.", "warning")
            return redirect(url_for("exams.final_exam_applications"))

        cur.execute(
            """
                UPDATE lms_final_exam_applications
                SET status = 'APPROVED'
                WHERE id = ?
            """,
            (application_id,),
        )
        log_activity(
            user_id=session.get("user_id"),
            branch_id=session.get("branch_id"),
            action_type="approve",
            module_name="lms_final_exam_applications",
            record_id=application_id,
            description=f"Approved final exam for {application['verified_name']}",
            conn=conn,
        )
        conn.commit()
        flash("Final exam application approved.", "success")
    finally:
        conn.close()

    return redirect(url_for("exams.final_exam_applications"))


@exams_bp.route("/lms_admin/final-exam/applications/<int:application_id>/reject", methods=["POST"])
@lms_content_manager_required
def reject_final_exam_application(application_id):
    conn = get_conn()
    try:
        cur = conn.cursor()
        application = _get_final_exam_application(cur, application_id)
        if not application:
            flash("Final exam application not found.", "warning")
            return redirect(url_for("exams.final_exam_applications"))

        cur.execute(
            """
                UPDATE lms_final_exam_applications
                SET status = 'REJECTED'
                WHERE id = ?
            """,
            (application_id,),
        )
        log_activity(
            user_id=session.get("user_id"),
            branch_id=session.get("branch_id"),
            action_type="reject",
            module_name="lms_final_exam_applications",
            record_id=application_id,
            description=f"Rejected final exam for {application['verified_name']}",
            conn=conn,
        )
        conn.commit()
        flash("Final exam application rejected.", "warning")
    finally:
        conn.close()

    return redirect(url_for("exams.final_exam_applications"))


@exams_bp.route("/student/final-exam/take/<int:application_id>", methods=["GET"])
def take_final_exam(application_id):
    auth_redirect = _student_required_redirect()
    if auth_redirect:
        return auth_redirect

    student_id = session["student_id"]
    conn = get_conn()
    try:
        cur = conn.cursor()
        application = _get_student_final_exam_application(cur, application_id, student_id)
        if not application:
            flash("Final exam application was not found.", "warning")
            return redirect(url_for("exams.final_exam_apply"))

        if not _is_final_exam_unlocked(application):
            flash("Your final exam is not unlocked yet.", "warning")
            return redirect(url_for("exams.final_exam_apply", program_id=application["course_id"]))

        attempt = _get_final_exam_attempt(cur, application_id, student_id)
        if attempt:
            return redirect(url_for("exams.final_exam_result", application_id=application_id))

        questions = list(_get_final_exam_questions(cur, application["course_id"]))
    finally:
        conn.close()

    if not questions:
        flash("No final exam questions are available for this program yet.", "warning")
        return redirect(url_for("exams.final_exam_apply", program_id=application["course_id"]))

    random.shuffle(questions)
    selected_questions = questions[:50]
    selected_ids = [question["id"] for question in selected_questions]

    conn = get_conn()
    try:
        answer_rows = conn.execute(
            f"""
                SELECT id, correct_option
                FROM lms_question_bank
                WHERE id IN ({','.join('?' for _ in selected_ids)})
            """,
            selected_ids,
        ).fetchall()
    finally:
        conn.close()

    session["final_exam_answers"] = {
        "application_id": application_id,
        "course_id": application["course_id"],
        "question_ids": selected_ids,
        "answers": {str(row["id"]): row["correct_option"] for row in answer_rows},
    }
    session.modified = True

    return render_template(
        "exams/take_exam.html",
        chapter=None,
        questions=selected_questions,
        question_count=len(selected_questions),
        exam_duration_seconds=60 * 60,
        exam_title="Final Exam",
        exam_context=application["program_name"] or application["course_name"] or "Approved Program",
        form_action=url_for("exams.submit_final_exam"),
        form_id="finalExamForm",
        submit_label="Submit Final Exam",
    )


@exams_bp.route("/student/final-exam/submit", methods=["POST"])
def submit_final_exam():
    auth_redirect = _student_required_redirect()
    if auth_redirect:
        return auth_redirect

    final_exam = session.get("final_exam_answers") or {}
    answer_key = final_exam.get("answers") or {}
    question_ids = final_exam.get("question_ids") or []
    application_id = final_exam.get("application_id")
    course_id = final_exam.get("course_id")

    if not answer_key or not question_ids or not application_id or not course_id:
        flash("Your final exam session has expired. Please start the exam again.", "warning")
        return redirect(url_for("exams.final_exam_apply"))

    submitted_answers = {
        str(question_id): request.form.get(f"answer_{question_id}", "").strip().upper()
        for question_id in question_ids
    }
    correct_count = sum(
        1
        for question_id, correct_option in answer_key.items()
        if submitted_answers.get(question_id) == correct_option
    )
    total_questions = len(question_ids)
    score_percent = round((correct_count / total_questions) * 100, 1) if total_questions else 0

    conn = get_conn()
    try:
        cur = conn.cursor()
        application = _get_student_final_exam_application(cur, application_id, session["student_id"])
        if not _is_final_exam_unlocked(application):
            flash("Your final exam is not unlocked yet.", "warning")
            return redirect(url_for("exams.final_exam_apply"))

        attempt = _get_final_exam_attempt(cur, application_id, session["student_id"])
        if not attempt:
            cur.execute(
                """
                    INSERT INTO lms_final_exam_attempts (
                        application_id,
                        student_id,
                        course_id,
                        question_ids_json,
                        submitted_answers_json,
                        correct_answers_json,
                        correct_count,
                        total_questions,
                        score_percent,
                        submitted_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                """,
                (
                    application_id,
                    session["student_id"],
                    course_id,
                    json.dumps(question_ids),
                    json.dumps(submitted_answers),
                    json.dumps(answer_key),
                    correct_count,
                    total_questions,
                    score_percent,
                ),
            )
            conn.commit()

            # Trigger certificate auto-issuance if settings allow and student is eligible
            try:
                from modules.certificates.services import EligibilityService, CertificateService
                settings = EligibilityService.get_settings(cur)
                if settings.get("auto_generate_certificates", 1) == 1 and score_percent >= settings.get("default_pass_percentage", 50.0):
                    CertificateService.issue_certificate(
                        conn, session["student_id"], course_id,
                        performed_by=None,
                        ip_address="Auto-Exam Submit",
                        user_agent="System Auto-Generation Flow"
                    )
                    conn.commit()
            except Exception as ex:
                # Log warning but do not block student submission flow
                print(f"Warning: Certificate auto-issuance skipped: {ex}")
                conn.rollback()
    finally:
        conn.close()

    session.pop("final_exam_answers", None)
    session.modified = True

    return redirect(url_for("exams.final_exam_result", application_id=application_id))


@exams_bp.route("/student/final-exam/result/<int:application_id>", methods=["GET"])
def final_exam_result(application_id):
    auth_redirect = _student_required_redirect()
    if auth_redirect:
        return auth_redirect

    conn = get_conn()
    try:
        cur = conn.cursor()
        application = _get_student_final_exam_application(cur, application_id, session["student_id"])
        attempt = _get_final_exam_attempt(cur, application_id, session["student_id"])
    finally:
        conn.close()

    if not application or not attempt:
        flash("Final exam result was not found.", "warning")
        return redirect(url_for("exams.final_exam_apply"))

    return render_template(
        "exams/mock_result.html",
        result_title="Final Exam Result",
        result_context=application["program_name"] or application["course_name"] or "Final Exam",
        correct_count=attempt["correct_count"],
        total_questions=attempt["total_questions"],
        score_percent=attempt["score_percent"],
        submitted_at=attempt["submitted_at"],
        review_items=[],
        back_url=url_for("exams.final_exam_apply", program_id=application["course_id"]),
        back_label="Back to Final Exam",
    )


@exams_bp.route("/student/mock/setup/<int:chapter_id>", methods=["GET"])
def chapter_mock_intro(chapter_id):
    auth_redirect = _student_required_redirect()
    if auth_redirect:
        return auth_redirect

    conn = get_conn()
    try:
        cur = conn.cursor()
        chapter = cur.execute(
            "SELECT id, title FROM lms_master_chapters WHERE id = ?",
            (chapter_id,),
        ).fetchone()
        if not chapter:
            flash("Chapter mock test was not found.", "warning")
            return redirect(url_for("students.dashboard"))

        questions = cur.execute(
            """
                SELECT
                    id,
                    chapter_id,
                    master_topic_id,
                    question_text,
                    option_a,
                    option_b,
                    option_c,
                    option_d,
                    question_type
                FROM lms_question_bank
                WHERE chapter_id = ?
            """,
            (chapter_id,),
        ).fetchall()
    finally:
        conn.close()

    questions = list(questions)
    if not questions:
        flash("No mock test questions are available for this chapter yet.", "warning")
        return redirect(url_for("students.dashboard"))

    # Group questions by master_topic_id (using None for no topic)
    grouped = {}
    for q in questions:
        tid = q["master_topic_id"]
        if tid not in grouped:
            grouped[tid] = []
        grouped[tid].append(dict(q))

    # Shuffle each group's questions
    for tid in grouped:
        random.shuffle(grouped[tid])

    # Round-robin distribution
    selected_questions = []
    keys = list(grouped.keys())
    # Shuffle keys to prevent topic order bias in round-robin
    random.shuffle(keys)

    while len(selected_questions) < 20 and any(grouped[k] for k in keys):
        for k in keys:
            if len(selected_questions) >= 20:
                break
            if grouped[k]:
                selected_questions.append(grouped[k].pop())

    # Shuffle the final selection so they are not ordered by topic round-robin
    random.shuffle(selected_questions)
    selected_ids = [question["id"] for question in selected_questions]

    conn = get_conn()
    try:
        answer_rows = conn.execute(
            f"""
                SELECT id, correct_option
                FROM lms_question_bank
                WHERE id IN ({','.join('?' for _ in selected_ids)})
            """,
            selected_ids,
        ).fetchall()
    finally:
        conn.close()

    session["chapter_mock_answers"] = {
        "chapter_id": chapter_id,
        "question_ids": selected_ids,
        "answers": {str(row["id"]): row["correct_option"] for row in answer_rows},
    }
    session.modified = True

    return render_template(
        "exams/take_exam.html",
        chapter=chapter,
        questions=selected_questions,
        question_count=len(selected_questions),
        exam_duration_seconds=20 * 60,
    )


def _build_mock_review_items(cur, question_ids, submitted_answers, correct_answers):
    if not question_ids:
        return []

    rows = cur.execute(
        f"""
            SELECT
                id,
                question_text,
                option_a,
                option_b,
                option_c,
                option_d
            FROM lms_question_bank
            WHERE id IN ({','.join('?' for _ in question_ids)})
        """,
        question_ids,
    ).fetchall()
    by_id = {row["id"]: row for row in rows}
    review_items = []

    for question_id in question_ids:
        question = by_id.get(question_id)
        if not question:
            continue
        question_key = str(question_id)
        student_answer = submitted_answers.get(question_key, "")
        correct_answer = correct_answers.get(question_key, "")
        review_items.append({
            "question": question,
            "student_answer": student_answer,
            "correct_answer": correct_answer,
            "is_correct": bool(student_answer and student_answer == correct_answer),
        })

    return review_items


@exams_bp.route("/student/mock/submit", methods=["POST"])
def submit_chapter_mock():
    auth_redirect = _student_required_redirect()
    if auth_redirect:
        return auth_redirect

    mock = session.get("chapter_mock_answers") or {}
    answer_key = mock.get("answers") or {}
    question_ids = mock.get("question_ids") or []
    chapter_id = mock.get("chapter_id")

    if not answer_key or not question_ids or not chapter_id:
        flash("Your mock test session has expired. Please start the mock test again.", "warning")
        return redirect(url_for("students.dashboard"))

    submitted_answers = {
        str(question_id): request.form.get(f"answer_{question_id}", "").strip().upper()
        for question_id in question_ids
    }
    correct_count = sum(
        1
        for question_id, correct_option in answer_key.items()
        if submitted_answers.get(question_id) == correct_option
    )
    total_questions = len(question_ids)
    score_percent = round((correct_count / total_questions) * 100, 1) if total_questions else 0

    conn = get_conn()
    try:
        cur = conn.cursor()
        chapter = cur.execute(
            "SELECT id, title FROM lms_master_chapters WHERE id = ?",
            (chapter_id,),
        ).fetchone()
        cur.execute(
            """
                INSERT INTO lms_chapter_mock_attempts (
                    student_id,
                    chapter_id,
                    question_ids_json,
                    submitted_answers_json,
                    correct_answers_json,
                    correct_count,
                    total_questions,
                    score_percent,
                    submitted_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (
                session["student_id"],
                chapter_id,
                json.dumps(question_ids),
                json.dumps(submitted_answers),
                json.dumps(answer_key),
                correct_count,
                total_questions,
                score_percent,
            ),
        )
        review_items = _build_mock_review_items(cur, question_ids, submitted_answers, answer_key)
        conn.commit()
    finally:
        conn.close()

    session.pop("chapter_mock_answers", None)
    session.modified = True

    return render_template(
        "exams/mock_result.html",
        chapter=chapter,
        correct_count=correct_count,
        total_questions=total_questions,
        score_percent=score_percent,
        review_items=review_items,
    )


@exams_bp.route("/student/mock/review/<int:chapter_id>", methods=["GET"])
def review_chapter_mock(chapter_id):
    auth_redirect = _student_required_redirect()
    if auth_redirect:
        return auth_redirect

    conn = get_conn()
    try:
        cur = conn.cursor()
        attempt = cur.execute(
            """
                SELECT *
                FROM lms_chapter_mock_attempts
                WHERE student_id = ?
                  AND chapter_id = ?
                ORDER BY submitted_at DESC, id DESC
                LIMIT 1
            """,
            (session["student_id"], chapter_id),
        ).fetchone()
        if not attempt:
            flash("Complete the chapter mock test first.", "warning")
            return redirect(url_for("exams.chapter_mock_intro", chapter_id=chapter_id))

        chapter = cur.execute(
            "SELECT id, title FROM lms_master_chapters WHERE id = ?",
            (chapter_id,),
        ).fetchone()

        question_ids = json.loads(attempt["question_ids_json"] or "[]")
        submitted_answers = json.loads(attempt["submitted_answers_json"] or "{}")
        correct_answers = json.loads(attempt["correct_answers_json"] or "{}")
        review_items = _build_mock_review_items(cur, question_ids, submitted_answers, correct_answers)
    finally:
        conn.close()

    return render_template(
        "exams/mock_result.html",
        chapter=chapter,
        correct_count=attempt["correct_count"],
        total_questions=attempt["total_questions"],
        score_percent=attempt["score_percent"],
        review_items=review_items,
        submitted_at=attempt["submitted_at"],
    )
