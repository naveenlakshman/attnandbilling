import io
import csv
import calendar
import openpyxl
import openpyxl.styles
from collections import defaultdict
from flask import Blueprint, render_template, send_file, flash, redirect, url_for, session, request, jsonify
from db import get_conn, log_activity
from modules.core.utils import login_required, admin_required
from werkzeug.security import generate_password_hash
from datetime import datetime, timedelta, timezone
from services.tenant_context import get_current_institute_id
from services.subscriptions import lock_and_check_limit

reports_bp = Blueprint("reports", __name__)


def parse_date(date_str):
    if not date_str:
        return None
    date_str = date_str.strip()
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass
    try:
        return datetime.strptime(date_str, "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        pass
    try:
        return datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        pass
    return None


@reports_bp.route("/")
@login_required
@admin_required
def dashboard():
    """Analytics and Reports Dashboard"""
    conn = get_conn()
    cur = conn.cursor()
    current_inst = get_current_institute_id(default=1)
    stats = {}
    tables = [
        ("branches", "Branches"),
        ("users", "Users"),
        ("leads", "Leads"),
        ("students", "Students"),
        ("courses", "Courses"),
        ("invoices", "Invoices"),
        ("receipts", "Receipts"),
        ("expenses", "Expenses"),
        ("expense_categories", "Expense Categories"),
        ("followups", "Followups"),
        ("installment_plans", "Installment Plans"),
        ("invoice_items", "Invoice Items"),
        ("activity_logs", "Activity Logs")
    ]
    
    for table_name, display_name in tables:
        try:
            if table_name in ("branches", "expenses", "invoices", "leads", "receipts", "students", "installment_plans", "activity_logs", "expense_categories"):
                cur.execute(f"SELECT COUNT(*) as count FROM {table_name} WHERE institute_id = ?", (current_inst,))
            elif table_name == "invoice_items":
                cur.execute("SELECT COUNT(*) as count FROM invoice_items ii JOIN invoices inv ON inv.id = ii.invoice_id WHERE inv.institute_id = ?", (current_inst,))
            elif table_name == "followups":
                cur.execute("SELECT COUNT(*) as count FROM followups f JOIN leads l ON l.id = f.lead_id WHERE l.institute_id = ?", (current_inst,))
            elif table_name == "users":
                cur.execute("SELECT COUNT(*) as count FROM users WHERE branch_id IN (SELECT id FROM branches WHERE institute_id = ?)", (current_inst,))
            else:
                cur.execute(f"SELECT COUNT(*) as count FROM {table_name}")
            result = cur.fetchone()
            stats[table_name] = {
                "name": display_name,
                "count": result["count"] if result else 0
            }
        except Exception as e:
            stats[table_name] = {
                "name": display_name,
                "count": 0,
                "error": str(e)
            }
    
    conn.close()
    return render_template("reports/dashboard.html", stats=stats)

    # ── 2a. Overdue Followups (past due, not yet done today) ─────
    cur.execute("""
        SELECT l.id, l.name, l.phone, l.stage, l.status, l.next_followup_date,
               u.full_name AS owner_name
        FROM leads l
        LEFT JOIN users u ON l.assigned_to_id = u.id
        WHERE l.next_followup_date < ? AND l.is_deleted = 0
          AND l.next_followup_date IS NOT NULL AND l.next_followup_date != ''
          AND l.status NOT IN ('converted', 'lost', 'not_interested')
        ORDER BY l.next_followup_date ASC, l.name
    """, (report_date,))
    followups_overdue = cur.fetchall()

    # ── 2b. Today's Followups done (actually logged today) ────────
    cur.execute("""
        SELECT f.id, f.method, f.outcome, f.note, f.created_at,
               l.id AS lead_id, l.name AS lead_name, l.phone AS lead_phone,
               u.full_name AS done_by
        FROM followups f
        JOIN leads l ON f.lead_id = l.id
        LEFT JOIN users u ON f.user_id = u.id
        WHERE substr(f.created_at, 1, 10) = ? AND l.is_deleted = 0
        ORDER BY f.created_at DESC
    """, (report_date,))
    followups_done = cur.fetchall()

    # ── 3. Today's Invoices ───────────────────────────────────────
    invoice_query = """
        SELECT i.id, i.invoice_no, i.invoice_date, i.total_amount, i.status,
               IFNULL((SELECT SUM(r2.amount_received) FROM receipts r2 WHERE r2.invoice_id = i.id), 0) AS paid_amount,
               (i.total_amount - IFNULL((SELECT SUM(r2.amount_received) FROM receipts r2 WHERE r2.invoice_id = i.id), 0)) AS balance_amount,
               s.full_name AS student_name, s.student_code, s.id AS student_id,
               br.branch_name
        FROM invoices i
        JOIN students s ON i.student_id = s.id
        LEFT JOIN branches br ON i.branch_id = br.id
        WHERE parse_date(i.invoice_date) = ?
    """
    invoice_params = [report_date]
    if selected_branch_id:
        invoice_query += " AND i.branch_id = ?"
        invoice_params.append(selected_branch_id)
    invoice_query += " ORDER BY i.created_at DESC"
    cur.execute(invoice_query, invoice_params)
    invoices = cur.fetchall()

    # ── 4. Today's Receipts ───────────────────────────────────────
    receipt_query = """
        SELECT r.id, r.receipt_no, r.receipt_date, r.amount_received, r.payment_mode,
               s.full_name AS student_name, s.student_code, i.invoice_no,
               br.branch_name
        FROM receipts r
        JOIN invoices i ON r.invoice_id = i.id
        JOIN students s ON i.student_id = s.id
        LEFT JOIN branches br ON i.branch_id = br.id
        WHERE parse_date(r.receipt_date) = ?
    """
    receipt_params = [report_date]
    if selected_branch_id:
        receipt_query += " AND i.branch_id = ?"
        receipt_params.append(selected_branch_id)
    receipt_query += " ORDER BY r.created_at DESC"
    cur.execute(receipt_query, receipt_params)
    receipts = cur.fetchall()

    # ── 5. Today's Attendance ─────────────────────────────────────
    att_summary = {"present": 0, "absent": 0, "late": 0, "leave": 0, "total": 0}
    att_records = []
    if selected_branch_id:
        cur.execute("""
            SELECT ar.status, COUNT(*) AS cnt
            FROM attendance_records ar
            WHERE ar.attendance_date = ? AND ar.branch_id = ?
            GROUP BY ar.status
        """, (report_date, selected_branch_id))
        for row in cur.fetchall():
            s = row["status"]
            if s in att_summary:
                att_summary[s] = row["cnt"]
            att_summary["total"] += row["cnt"]

        cur.execute("""
            SELECT ar.status, s.full_name AS student_name, s.student_code,
                   b.batch_name, br2.branch_name
            FROM attendance_records ar
            JOIN students s ON ar.student_id = s.id
            JOIN batches b ON ar.batch_id = b.id
            LEFT JOIN branches br2 ON ar.branch_id = br2.id
            WHERE ar.attendance_date = ? AND ar.branch_id = ?
            ORDER BY b.batch_name, s.full_name
        """, (report_date, selected_branch_id))
        att_records = cur.fetchall()
    else:
        cur.execute("""
            SELECT ar.status, COUNT(*) AS cnt
            FROM attendance_records ar
            WHERE ar.attendance_date = ?
            GROUP BY ar.status
        """, (report_date,))
        for row in cur.fetchall():
            s = row["status"]
            if s in att_summary:
                att_summary[s] = row["cnt"]
            att_summary["total"] += row["cnt"]

        cur.execute("""
            SELECT ar.status, s.full_name AS student_name, s.student_code,
                   b.batch_name, br2.branch_name
            FROM attendance_records ar
            JOIN students s ON ar.student_id = s.id
            JOIN batches b ON ar.batch_id = b.id
            LEFT JOIN branches br2 ON ar.branch_id = br2.id
            WHERE ar.attendance_date = ?
            ORDER BY br2.branch_name, b.batch_name, s.full_name
        """, (report_date,))
        att_records = cur.fetchall()

    conn.close()

    # Summary totals
    totals = {
        "new_leads": len(new_leads),
        "followups": len(followups_due),
        "followups_done": len(followups_done),
        "invoices": len(invoices),
        "invoice_amount": sum(r["total_amount"] or 0 for r in invoices),
        "receipts": len(receipts),
        "receipt_amount": sum(r["amount_received"] or 0 for r in receipts),
        "attendance": att_summary["total"],
    }

    return render_template(
        "reports/daily.html",
        report_date=report_date,
        branches=branches,
        selected_branch_id=selected_branch_id,
        can_view_all=can_view_all,
        new_leads=new_leads,
        followups_due=followups_due,
        followups_overdue=followups_overdue,
        followups_done=followups_done,
        invoices=invoices,
        receipts=receipts,
        att_summary=att_summary,
        att_records=att_records,
        totals=totals,
    )


@reports_bp.route("/daily/download")
@login_required
def daily_report_download():
    """Download daily report as CSV (all sections)"""
    IST = timezone(timedelta(hours=5, minutes=30))
    today_default = datetime.now(IST).strftime("%Y-%m-%d")

    report_date = request.args.get("date", today_default).strip()
    try:
        datetime.strptime(report_date, "%Y-%m-%d")
    except ValueError:
        report_date = today_default

    conn = get_conn()
    cur = conn.cursor()

    can_view_all = session.get("can_view_all_branches", False) or session.get("role") == "admin"
    user_branch_id = session.get("branch_id")
    branch_param = request.args.get("branch_id", "").strip()
    if can_view_all:
        if branch_param in ("", "all") or not branch_param:
            selected_branch_id = None
        else:
            selected_branch_id = int(branch_param) if branch_param.isdigit() else None
    else:
        selected_branch_id = user_branch_id

    # ── Branch name for filename ──────────────────────────────────
    branch_label = ""
    if selected_branch_id:
        cur.execute("SELECT branch_name FROM branches WHERE id = ?", (selected_branch_id,))
        br = cur.fetchone()
        if br:
            branch_label = "_" + br["branch_name"].replace(" ", "_")

    # ── New Leads ─────────────────────────────────────────────────
    cur.execute("""
        SELECT l.id, l.name, l.phone, l.lead_source, l.stage, l.status, l.created_at,
               u.full_name AS owner_name
        FROM leads l
        LEFT JOIN users u ON l.assigned_to_id = u.id
        WHERE substr(l.created_at, 1, 10) = ? AND l.is_deleted = 0
        ORDER BY l.created_at DESC
    """, (report_date,))
    new_leads = cur.fetchall()

    # ── Followups (due) ───────────────────────────────────────────
    cur.execute("""
        SELECT l.id, l.name, l.phone, l.stage, l.status, l.next_followup_date,
               u.full_name AS owner_name
        FROM leads l
        LEFT JOIN users u ON l.assigned_to_id = u.id
        WHERE l.next_followup_date = ? AND l.is_deleted = 0
          AND l.status NOT IN ('converted', 'lost', 'not_interested')
        ORDER BY l.name
    """, (report_date,))
    followups_due = cur.fetchall()

    # ── Followups done today ──────────────────────────────────────
    cur.execute("""
        SELECT f.id, f.method, f.outcome, f.note, f.created_at,
               l.name AS lead_name, l.phone AS lead_phone,
               u.full_name AS done_by
        FROM followups f
        JOIN leads l ON f.lead_id = l.id
        LEFT JOIN users u ON f.user_id = u.id
        WHERE substr(f.created_at, 1, 10) = ? AND l.is_deleted = 0
        ORDER BY f.created_at DESC
    """, (report_date,))
    followups_done = cur.fetchall()

    # ── Invoices ──────────────────────────────────────────────────
    invoice_query = """
        SELECT i.invoice_no, i.invoice_date, s.full_name AS student_name, s.student_code,
               i.total_amount,
               IFNULL((SELECT SUM(r2.amount_received) FROM receipts r2 WHERE r2.invoice_id = i.id), 0) AS paid_amount,
               (i.total_amount - IFNULL((SELECT SUM(r2.amount_received) FROM receipts r2 WHERE r2.invoice_id = i.id), 0)) AS balance_amount,
               i.status, br.branch_name
        FROM invoices i
        JOIN students s ON i.student_id = s.id
        LEFT JOIN branches br ON i.branch_id = br.id
        WHERE parse_date(i.invoice_date) = ?
    """
    invoice_params = [report_date]
    if selected_branch_id:
        invoice_query += " AND i.branch_id = ?"
        invoice_params.append(selected_branch_id)
    invoice_query += " ORDER BY i.created_at DESC"
    cur.execute(invoice_query, invoice_params)
    invoices = cur.fetchall()

    # ── Receipts ──────────────────────────────────────────────────
    receipt_query = """
        SELECT r.receipt_no, r.receipt_date, s.full_name AS student_name, s.student_code,
               i.invoice_no, r.amount_received, r.payment_mode, br.branch_name
        FROM receipts r
        JOIN invoices i ON r.invoice_id = i.id
        JOIN students s ON i.student_id = s.id
        LEFT JOIN branches br ON i.branch_id = br.id
        WHERE parse_date(r.receipt_date) = ?
    """
    receipt_params = [report_date]
    if selected_branch_id:
        receipt_query += " AND i.branch_id = ?"
        receipt_params.append(selected_branch_id)
    receipt_query += " ORDER BY r.created_at DESC"
    cur.execute(receipt_query, receipt_params)
    receipts = cur.fetchall()

    # ── Attendance ────────────────────────────────────────────────
    att_records = []
    if selected_branch_id:
        cur.execute("""
            SELECT s.full_name AS student_name, s.student_code,
                   b.batch_name, ar.status, br2.branch_name
            FROM attendance_records ar
            JOIN students s ON ar.student_id = s.id
            JOIN batches b ON ar.batch_id = b.id
            LEFT JOIN branches br2 ON ar.branch_id = br2.id
            WHERE ar.attendance_date = ? AND ar.branch_id = ?
            ORDER BY b.batch_name, s.full_name
        """, (report_date, selected_branch_id))
        att_records = cur.fetchall()
    else:
        cur.execute("""
            SELECT s.full_name AS student_name, s.student_code,
                   b.batch_name, ar.status, br2.branch_name
            FROM attendance_records ar
            JOIN students s ON ar.student_id = s.id
            JOIN batches b ON ar.batch_id = b.id
            LEFT JOIN branches br2 ON ar.branch_id = br2.id
            WHERE ar.attendance_date = ?
            ORDER BY br2.branch_name, b.batch_name, s.full_name
        """, (report_date,))
        att_records = cur.fetchall()

    conn.close()

    # ── Build CSV ─────────────────────────────────────────────────
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([f"Daily Report – {report_date}{(' – ' + br['branch_name']) if selected_branch_id and br else ''}"])
    writer.writerow([])

    # Section 1: New Leads
    writer.writerow(["NEW LEADS"])
    writer.writerow(["#", "Name", "Phone", "Source", "Stage", "Status", "Time"])
    for i, l in enumerate(new_leads, 1):
        writer.writerow([i, l["name"], l["phone"] or "", l["lead_source"] or "",
                         l["stage"] or "", l["status"] or "",
                         (l["created_at"] or "")[11:16]])
    if not new_leads:
        writer.writerow(["No new leads"])
    writer.writerow([])

    # Section 2: Followups
    writer.writerow(["FOLLOWUPS DUE"])
    writer.writerow(["#", "Name", "Phone", "Stage", "Status", "Due Date", "Owner"])
    for i, f in enumerate(followups_due, 1):
        writer.writerow([i, f["name"], f["phone"] or "", f["stage"] or "", f["status"] or "",
                         f["next_followup_date"] or "", f["owner_name"] or ""])
    if not followups_due:
        writer.writerow(["No followups today"])
    writer.writerow([])

    # Section 2b: Followups Done Today
    writer.writerow(["FOLLOWUPS DONE TODAY"])
    writer.writerow(["#", "Lead Name", "Phone", "Method", "Outcome", "Note", "Done By", "Time (IST)"])
    for i, f in enumerate(followups_done, 1):
        from datetime import datetime as _dt, timedelta as _td
        try:
            t = _dt.fromisoformat(f["created_at"]) + _td(hours=5, minutes=30)
            time_str = t.strftime("%I:%M %p")
        except Exception:
            time_str = (f["created_at"] or "")[11:16]
        writer.writerow([i, f["lead_name"], f["lead_phone"] or "",
                         f["method"] or "", f["outcome"] or "",
                         f["note"] or "", f["done_by"] or "", time_str])
    if not followups_done:
        writer.writerow(["No followups logged today"])
    writer.writerow([])

    # Section 3: Invoices
    writer.writerow(["INVOICES"])
    writer.writerow(["#", "Invoice No.", "Date", "Student", "Reg. No", "Total", "Paid", "Balance", "Status", "Branch"])
    inv_total = inv_paid = inv_balance = 0
    for i, inv in enumerate(invoices, 1):
        writer.writerow([i, inv["invoice_no"], inv["invoice_date"], inv["student_name"],
                         inv["student_code"], inv["total_amount"] or 0,
                         inv["paid_amount"] or 0, inv["balance_amount"] or 0,
                         inv["status"] or "", inv["branch_name"] or ""])
        inv_total += inv["total_amount"] or 0
        inv_paid += inv["paid_amount"] or 0
        inv_balance += inv["balance_amount"] or 0
    if not invoices:
        writer.writerow(["No invoices today"])
    else:
        writer.writerow(["", "", "", "", "TOTAL", inv_total, inv_paid, inv_balance, "", ""])
    writer.writerow([])

    # Section 4: Receipts
    writer.writerow(["RECEIPTS"])
    writer.writerow(["#", "Receipt No.", "Date", "Student", "Reg. No", "Invoice No.", "Amount", "Mode", "Branch"])
    rec_total = 0
    for i, r in enumerate(receipts, 1):
        writer.writerow([i, r["receipt_no"], r["receipt_date"], r["student_name"],
                         r["student_code"], r["invoice_no"],
                         r["amount_received"] or 0, r["payment_mode"] or "", r["branch_name"] or ""])
        rec_total += r["amount_received"] or 0
    if not receipts:
        writer.writerow(["No receipts today"])
    else:
        writer.writerow(["", "", "", "", "", "TOTAL", rec_total, "", ""])
    writer.writerow([])

    # Section 5: Attendance
    writer.writerow(["ATTENDANCE"])
    if att_records:
        if selected_branch_id:
            writer.writerow(["#", "Student", "Reg. No", "Batch", "Status"])
            for i, a in enumerate(att_records, 1):
                writer.writerow([i, a["student_name"], a["student_code"], a["batch_name"], a["status"]])
        else:
            writer.writerow(["#", "Student", "Reg. No", "Batch", "Branch", "Status"])
            for i, a in enumerate(att_records, 1):
                writer.writerow([i, a["student_name"], a["student_code"], a["batch_name"], a["branch_name"] or "", a["status"]])
    else:
        writer.writerow(["No attendance recorded today"])

    csv_data = output.getvalue()
    output.close()

    buf = io.BytesIO()
    buf.write(csv_data.encode("utf-8-sig"))  # utf-8-sig adds BOM for Excel compatibility
    buf.seek(0)

    filename = f"daily_report_{report_date}{branch_label}.csv"
    return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)


def _resolve_report_branch(cur):
    """Return branch filter details using the same access rules as daily reports."""
    current_inst = get_current_institute_id(default=1)
    cur.execute(
        """SELECT id, branch_name, branch_code
           FROM branches
           WHERE is_active = 1 AND institute_id = ?
           ORDER BY branch_name""",
        (current_inst,),
    )
    branches = cur.fetchall()

    can_view_all = session.get("can_view_all_branches", False) or session.get("role") == "admin"
    user_branch_id = session.get("branch_id")
    branch_param = request.args.get("branch_id", "").strip()

    allowed_branch_ids = {branch["id"] for branch in branches}
    if can_view_all:
        requested_branch_id = int(branch_param) if branch_param.isdigit() else None
        selected_branch_id = requested_branch_id if requested_branch_id in allowed_branch_ids else None
    else:
        selected_branch_id = user_branch_id if user_branch_id in allowed_branch_ids else None

    selected_branch_name = "All Branches"
    if selected_branch_id:
        for branch in branches:
            if branch["id"] == selected_branch_id:
                selected_branch_name = branch["branch_name"]
                break

    return branches, selected_branch_id, selected_branch_name, can_view_all


REPORT_IST = timezone(timedelta(hours=5, minutes=30))


def _today_ist():
    return datetime.now(REPORT_IST).strftime("%Y-%m-%d")


def _month_bounds(month_value):
    """Validate YYYY-MM and return month metadata for queries and headings."""
    today_month = datetime.now(REPORT_IST).strftime("%Y-%m")

    try:
        month_start = datetime.strptime(month_value or today_month, "%Y-%m")
    except ValueError:
        month_start = datetime.strptime(today_month, "%Y-%m")

    last_day = calendar.monthrange(month_start.year, month_start.month)[1]
    start_date = month_start.strftime("%Y-%m-01")
    end_date = month_start.replace(day=last_day).strftime("%Y-%m-%d")
    month_label = month_start.strftime("%B %Y")
    month_value = month_start.strftime("%Y-%m")

    return month_value, month_label, start_date, end_date, last_day


def _calculation_window(start_date, end_date):
    """Never calculate expected/unmarked attendance for future dates."""
    today = _today_ist()
    calculation_end = min(end_date, today)
    if start_date > calculation_end:
        query_end = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        return today, calculation_end, query_end, False
    return today, calculation_end, calculation_end, True


def _attendance_where(selected_branch_id):
    where = [
        "ar.attendance_date BETWEEN ? AND ?",
        "EXISTS (SELECT 1 FROM branches tenant_branch "
        "WHERE tenant_branch.id = ar.branch_id AND tenant_branch.institute_id = ?)",
    ]
    if selected_branch_id:
        where.append("ar.branch_id = ?")
    return " AND ".join(where)


def _attendance_params(start_date, end_date, selected_branch_id):
    params = [start_date, end_date, get_current_institute_id(default=1)]
    if selected_branch_id:
        params.append(selected_branch_id)
    return params


def _date_range(start_date, end_date):
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    days = []
    while start <= end:
        days.append(start.strftime("%Y-%m-%d"))
        start += timedelta(days=1)
    return days


def _ensure_attendance_calendar_tables(cur):
    now = datetime.now(REPORT_IST).strftime("%Y-%m-%d %H:%M:%S")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS tenant_attendance_calendar_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            institute_id INTEGER NOT NULL UNIQUE,
            working_days TEXT NOT NULL DEFAULT '0,1,2,3,4,5',
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS tenant_attendance_holidays (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            institute_id INTEGER NOT NULL,
            holiday_date TEXT NOT NULL,
            title TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT,
            UNIQUE(institute_id, holiday_date)
        )
    """)
    current_inst = get_current_institute_id(default=1)
    cur.execute(
        "SELECT id FROM tenant_attendance_calendar_settings WHERE institute_id = ?",
        (current_inst,),
    )
    if not cur.fetchone():
        cur.execute("""
            INSERT INTO tenant_attendance_calendar_settings
                (institute_id, working_days, created_at)
            VALUES (?, '0,1,2,3,4,5', ?)
        """, (current_inst, now))


def _get_attendance_calendar(cur, start_date=None, end_date=None):
    _ensure_attendance_calendar_tables(cur)
    current_inst = get_current_institute_id(default=1)
    row = cur.execute(
        "SELECT working_days FROM tenant_attendance_calendar_settings WHERE institute_id = ?",
        (current_inst,),
    ).fetchone()
    raw_days = (row["working_days"] if row else "0,1,2,3,4,5") or "0,1,2,3,4,5"
    working_days = {
        int(day)
        for day in raw_days.split(",")
        if day.strip().isdigit() and 0 <= int(day.strip()) <= 6
    }
    if not working_days:
        working_days = {0, 1, 2, 3, 4, 5}

    if start_date and end_date:
        cur.execute("""
            SELECT id, holiday_date, title
            FROM tenant_attendance_holidays
            WHERE institute_id = ? AND holiday_date BETWEEN ? AND ?
            ORDER BY holiday_date
        """, (current_inst, start_date, end_date))
    else:
        cur.execute("""
            SELECT id, holiday_date, title
            FROM tenant_attendance_holidays
            WHERE institute_id = ?
            ORDER BY holiday_date DESC
        """, (current_inst,))
    holidays = [dict(row) for row in cur.fetchall()]
    holiday_map = {row["holiday_date"]: row["title"] for row in holidays}

    return {
        "working_days": working_days,
        "working_days_csv": ",".join(str(day) for day in sorted(working_days)),
        "holidays": holidays,
        "holiday_map": holiday_map,
    }


def _is_expected_working_date(date_value, calendar_settings):
    date_obj = datetime.strptime(date_value, "%Y-%m-%d")
    if date_obj.weekday() not in calendar_settings["working_days"]:
        return False
    return date_value not in calendar_settings["holiday_map"]


def _clamp_date(value, fallback):
    if not value:
        return fallback
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return fallback


def _load_expected_attendance(cur, start_date, end_date, selected_branch_id, calendar_settings):
    """Build expected student/batch attendance rows for each day in a report period."""
    query = """
        SELECT
            sb.student_id,
            sb.batch_id,
            sb.joined_on,
            s.student_code,
            s.full_name,
            s.phone,
            b.batch_name,
            b.branch_id,
            b.start_date,
            b.end_date,
            c.course_name,
            br.branch_name,
            u.full_name AS trainer_name,
            b.trainer_id
        FROM student_batches sb
        JOIN students s ON sb.student_id = s.id
        JOIN batches b ON sb.batch_id = b.id
        LEFT JOIN courses c ON b.course_id = c.id
        LEFT JOIN branches br ON b.branch_id = br.id
        LEFT JOIN users u ON b.trainer_id = u.id
        WHERE sb.status = 'active'
          AND b.status = 'active'
          AND s.institute_id = ?
          AND br.institute_id = ?
          AND (b.start_date IS NULL OR date(b.start_date) <= date(?))
          AND (b.end_date IS NULL OR date(b.end_date) >= date(?))
    """
    current_inst = get_current_institute_id(default=1)
    params = [current_inst, current_inst, end_date, start_date]
    if selected_branch_id:
        query += " AND b.branch_id = ?"
        params.append(selected_branch_id)
    query += " ORDER BY br.branch_name, b.batch_name, s.full_name"

    cur.execute(query, params)
    roster_rows = [dict(row) for row in cur.fetchall()]

    period_dates = [
        day for day in _date_range(start_date, end_date)
        if _is_expected_working_date(day, calendar_settings)
    ]
    expected_keys = set()
    expected_by_day = defaultdict(set)
    expected_by_branch = defaultdict(set)
    expected_by_batch = defaultdict(set)
    expected_by_student_batch = defaultdict(set)
    expected_row_lookup = {}

    for row in roster_rows:
        effective_start = max(
            start_date,
            _clamp_date(row.get("start_date"), start_date),
            _clamp_date(row.get("joined_on"), start_date),
        )
        effective_end = min(end_date, _clamp_date(row.get("end_date"), end_date))
        if effective_start > effective_end:
            continue

        branch_key = row.get("branch_id") or 0
        batch_key = row.get("batch_id") or 0
        student_batch_key = (row.get("student_id"), batch_key)
        expected_row_lookup[student_batch_key] = row

        for day in period_dates:
            if day < effective_start or day > effective_end:
                continue
            key = (day, row.get("student_id"), batch_key)
            expected_keys.add(key)
            expected_by_day[day].add(key)
            expected_by_branch[branch_key].add(key)
            expected_by_batch[batch_key].add(key)
            expected_by_student_batch[student_batch_key].add(key)

    return {
        "period_dates": period_dates,
        "keys": expected_keys,
        "by_day": expected_by_day,
        "by_branch": expected_by_branch,
        "by_batch": expected_by_batch,
        "by_student_batch": expected_by_student_batch,
        "row_lookup": expected_row_lookup,
    }


def _load_marked_keys(cur, start_date, end_date, selected_branch_id):
    where_clause = _attendance_where(selected_branch_id)
    params = _attendance_params(start_date, end_date, selected_branch_id)
    cur.execute(f"""
        SELECT ar.attendance_date, ar.student_id, ar.batch_id
        FROM attendance_records ar
        WHERE {where_clause}
    """, params)
    return {
        (row["attendance_date"], row["student_id"], row["batch_id"] or 0)
        for row in cur.fetchall()
    }


@reports_bp.route("/attendance/monthly")
@login_required
def attendance_monthly_report():
    """Month-level attendance report for the Reports module."""
    report_month, month_label, start_date, end_date, last_day = _month_bounds(request.args.get("month"))
    today, calculation_end_date, query_end_date, has_calculation_window = _calculation_window(start_date, end_date)

    conn = get_conn()
    cur = conn.cursor()
    try:
        branches, selected_branch_id, selected_branch_name, can_view_all = _resolve_report_branch(cur)
        calendar_settings = _get_attendance_calendar(cur, start_date, end_date)
        where_clause = _attendance_where(selected_branch_id)
        params = _attendance_params(start_date, query_end_date, selected_branch_id)
        expected_data = _load_expected_attendance(cur, start_date, query_end_date, selected_branch_id, calendar_settings)
        marked_keys = _load_marked_keys(cur, start_date, query_end_date, selected_branch_id)

        cur.execute(f"""
            SELECT
                COUNT(*) AS total_marked,
                COUNT(DISTINCT ar.student_id) AS unique_students,
                COUNT(DISTINCT ar.batch_id) AS unique_batches,
                SUM(CASE WHEN ar.status = 'present' THEN 1 ELSE 0 END) AS present,
                SUM(CASE WHEN ar.status = 'absent' THEN 1 ELSE 0 END) AS absent,
                SUM(CASE WHEN ar.status = 'late' THEN 1 ELSE 0 END) AS late,
                SUM(CASE WHEN ar.status = 'leave' THEN 1 ELSE 0 END) AS leave_count
            FROM attendance_records ar
            WHERE {where_clause}
        """, params)
        raw_totals = cur.fetchone()
        totals = {
            "total_marked": raw_totals["total_marked"] or 0,
            "unique_students": raw_totals["unique_students"] or 0,
            "unique_batches": raw_totals["unique_batches"] or 0,
            "present": raw_totals["present"] or 0,
            "absent": raw_totals["absent"] or 0,
            "late": raw_totals["late"] or 0,
            "leave": raw_totals["leave_count"] or 0,
        }
        attended = totals["present"] + totals["late"]
        totals["attendance_rate"] = round((attended / totals["total_marked"] * 100), 1) if totals["total_marked"] else 0
        totals["expected_records"] = len(expected_data["keys"])
        totals["expected_unique_students"] = len({key[1] for key in expected_data["keys"]})
        totals["expected_unique_batches"] = len({key[2] for key in expected_data["keys"]})
        totals["unmarked"] = 0
        totals["marking_completion_rate"] = 0

        cur.execute(f"""
            SELECT
                ar.attendance_date,
                COUNT(*) AS total_marked,
                SUM(CASE WHEN ar.status = 'present' THEN 1 ELSE 0 END) AS present,
                SUM(CASE WHEN ar.status = 'absent' THEN 1 ELSE 0 END) AS absent,
                SUM(CASE WHEN ar.status = 'late' THEN 1 ELSE 0 END) AS late,
                SUM(CASE WHEN ar.status = 'leave' THEN 1 ELSE 0 END) AS leave_count
            FROM attendance_records ar
            WHERE {where_clause}
            GROUP BY ar.attendance_date
            ORDER BY ar.attendance_date
        """, params)
        daily_map = {row["attendance_date"]: dict(row) for row in cur.fetchall()}
        daily_breakdown = []
        monthly_unmarked = 0
        for day in range(1, last_day + 1):
            day_date = f"{report_month}-{day:02d}"
            row = daily_map.get(day_date, {})
            total_marked = row.get("total_marked") or 0
            present = row.get("present") or 0
            late = row.get("late") or 0
            is_future = day_date > today
            is_working_day = _is_expected_working_date(day_date, calendar_settings)
            holiday_title = calendar_settings["holiday_map"].get(day_date)
            day_expected_keys = expected_data["by_day"].get(day_date, set())
            expected_for_day = len(day_expected_keys)
            day_unmarked = sum(1 for key in day_expected_keys if key not in marked_keys)
            monthly_unmarked += day_unmarked
            daily_breakdown.append({
                "attendance_date": day_date,
                "day_name": datetime.strptime(day_date, "%Y-%m-%d").strftime("%a"),
                "expected": expected_for_day,
                "total_marked": total_marked,
                "unmarked": day_unmarked,
                "present": present,
                "absent": row.get("absent") or 0,
                "late": late,
                "leave": row.get("leave_count") or 0,
                "rate": round(((present + late) / total_marked * 100), 1) if total_marked else 0,
                "is_future": is_future,
                "is_working_day": is_working_day,
                "holiday_title": holiday_title,
                "calendar_note": (
                    "Future"
                    if is_future else
                    holiday_title
                    if holiday_title else
                    "Weekly Off"
                    if not is_working_day else
                    ""
                ),
            })
        totals["unmarked"] = monthly_unmarked
        totals["marking_completion_rate"] = (
            round(((totals["expected_records"] - monthly_unmarked) / totals["expected_records"] * 100), 1)
            if totals["expected_records"] else 0
        )

        cur.execute(f"""
            SELECT
                ar.branch_id,
                br.branch_name,
                COUNT(*) AS total_marked,
                COUNT(DISTINCT ar.student_id) AS unique_students,
                COUNT(DISTINCT ar.batch_id) AS unique_batches,
                SUM(CASE WHEN ar.status = 'present' THEN 1 ELSE 0 END) AS present,
                SUM(CASE WHEN ar.status = 'absent' THEN 1 ELSE 0 END) AS absent,
                SUM(CASE WHEN ar.status = 'late' THEN 1 ELSE 0 END) AS late,
                SUM(CASE WHEN ar.status = 'leave' THEN 1 ELSE 0 END) AS leave_count
            FROM attendance_records ar
            LEFT JOIN branches br ON ar.branch_id = br.id
            WHERE {where_clause}
            GROUP BY ar.branch_id, br.branch_name
            ORDER BY br.branch_name
        """, params)
        branch_summary = []
        for row in cur.fetchall():
            total_marked = row["total_marked"] or 0
            present = row["present"] or 0
            late = row["late"] or 0
            branch_expected_keys = expected_data["by_branch"].get(row["branch_id"] or 0, set())
            expected = len(branch_expected_keys)
            branch_summary.append({
                **dict(row),
                "unique_students": max(row["unique_students"] or 0, len({key[1] for key in branch_expected_keys})),
                "unique_batches": max(row["unique_batches"] or 0, len({key[2] for key in branch_expected_keys})),
                "expected": expected,
                "unmarked": sum(1 for key in branch_expected_keys if key not in marked_keys),
                "leave": row["leave_count"] or 0,
                "rate": round(((present + late) / total_marked * 100), 1) if total_marked else 0,
                "marking_rate": round((min(total_marked, expected) / expected * 100), 1) if expected else 0,
            })
        branch_ids_with_rows = {row["branch_id"] or 0 for row in branch_summary}
        branch_lookup = {branch["id"]: branch["branch_name"] for branch in branches}
        for branch_id, keys in expected_data["by_branch"].items():
            if branch_id in branch_ids_with_rows:
                continue
            expected = len(keys)
            branch_summary.append({
                "branch_id": branch_id,
                "branch_name": branch_lookup.get(branch_id, "Unassigned"),
                "total_marked": 0,
                "unique_students": len({key[1] for key in keys}),
                "unique_batches": len({key[2] for key in keys}),
                "present": 0,
                "absent": 0,
                "late": 0,
                "leave_count": 0,
                "expected": expected,
                "unmarked": expected,
                "leave": 0,
                "rate": 0,
                "marking_rate": 0,
            })
        branch_summary.sort(key=lambda row: row.get("branch_name") or "")

        cur.execute(f"""
            SELECT
                b.id AS batch_id,
                b.batch_name,
                c.course_name,
                br.branch_name,
                u.full_name AS trainer_name,
                COUNT(*) AS total_marked,
                COUNT(DISTINCT ar.student_id) AS unique_students,
                SUM(CASE WHEN ar.status = 'present' THEN 1 ELSE 0 END) AS present,
                SUM(CASE WHEN ar.status = 'absent' THEN 1 ELSE 0 END) AS absent,
                SUM(CASE WHEN ar.status = 'late' THEN 1 ELSE 0 END) AS late,
                SUM(CASE WHEN ar.status = 'leave' THEN 1 ELSE 0 END) AS leave_count
            FROM attendance_records ar
            LEFT JOIN batches b ON ar.batch_id = b.id
            LEFT JOIN courses c ON b.course_id = c.id
            LEFT JOIN branches br ON ar.branch_id = br.id
            LEFT JOIN users u ON b.trainer_id = u.id
            WHERE {where_clause}
            GROUP BY ar.batch_id, b.batch_name, c.course_name, br.branch_name, u.full_name
            ORDER BY br.branch_name, b.batch_name
        """, params)
        batch_summary = []
        for row in cur.fetchall():
            total_marked = row["total_marked"] or 0
            present = row["present"] or 0
            late = row["late"] or 0
            batch_expected_keys = expected_data["by_batch"].get(row["batch_id"] or 0, set())
            expected = len(batch_expected_keys)
            batch_summary.append({
                **dict(row),
                "unique_students": max(row["unique_students"] or 0, len({key[1] for key in batch_expected_keys})),
                "expected": expected,
                "unmarked": sum(1 for key in batch_expected_keys if key not in marked_keys),
                "leave": row["leave_count"] or 0,
                "rate": round(((present + late) / total_marked * 100), 1) if total_marked else 0,
                "marking_rate": round((min(total_marked, expected) / expected * 100), 1) if expected else 0,
            })
        batch_ids_with_rows = {row["batch_id"] or 0 for row in batch_summary}
        for batch_id, keys in expected_data["by_batch"].items():
            if batch_id in batch_ids_with_rows:
                continue
            sample_row = next((row for row in expected_data["row_lookup"].values() if (row.get("batch_id") or 0) == batch_id), {})
            expected = len(keys)
            batch_summary.append({
                "batch_id": batch_id,
                "batch_name": sample_row.get("batch_name") or "Batch",
                "course_name": sample_row.get("course_name"),
                "branch_name": sample_row.get("branch_name"),
                "trainer_name": sample_row.get("trainer_name"),
                "total_marked": 0,
                "unique_students": len({key[1] for key in keys}),
                "present": 0,
                "absent": 0,
                "late": 0,
                "leave_count": 0,
                "expected": expected,
                "unmarked": expected,
                "leave": 0,
                "rate": 0,
                "marking_rate": 0,
            })
        batch_summary.sort(key=lambda row: ((row.get("branch_name") or ""), (row.get("batch_name") or "")))

        cur.execute(f"""
            SELECT
                s.id AS student_id,
                s.student_code,
                s.full_name,
                s.phone,
                b.id AS batch_id,
                b.batch_name,
                br.id AS branch_id,
                br.branch_name,
                COUNT(*) AS total_marked,
                SUM(CASE WHEN ar.status = 'present' THEN 1 ELSE 0 END) AS present,
                SUM(CASE WHEN ar.status = 'absent' THEN 1 ELSE 0 END) AS absent,
                SUM(CASE WHEN ar.status = 'late' THEN 1 ELSE 0 END) AS late,
                SUM(CASE WHEN ar.status = 'leave' THEN 1 ELSE 0 END) AS leave_count,
                MAX(ar.attendance_date) AS last_marked
            FROM attendance_records ar
            JOIN students s ON ar.student_id = s.id
            LEFT JOIN batches b ON ar.batch_id = b.id
            LEFT JOIN branches br ON ar.branch_id = br.id
            WHERE {where_clause}
            GROUP BY ar.student_id, ar.batch_id, s.student_code, s.full_name, s.phone, b.id, b.batch_name, br.id, br.branch_name
            ORDER BY s.full_name, b.batch_name
        """, params)
        student_summary = []
        for row in cur.fetchall():
            total_marked = row["total_marked"] or 0
            present = row["present"] or 0
            late = row["late"] or 0
            student_batch_key = (row["student_id"], row["batch_id"] or 0)
            student_expected_keys = expected_data["by_student_batch"].get(student_batch_key, set())
            expected = len(student_expected_keys)
            student_summary.append({
                **dict(row),
                "expected": expected,
                "unmarked": sum(1 for key in student_expected_keys if key not in marked_keys),
                "leave": row["leave_count"] or 0,
                "rate": round(((present + late) / total_marked * 100), 1) if total_marked else 0,
                "marking_rate": round((min(total_marked, expected) / expected * 100), 1) if expected else 0,
            })
        student_batch_keys_with_rows = {(row["student_id"], row["batch_id"] or 0) for row in student_summary}
        for student_batch_key, keys in expected_data["by_student_batch"].items():
            if student_batch_key in student_batch_keys_with_rows:
                continue
            sample_row = expected_data["row_lookup"].get(student_batch_key, {})
            expected = len(keys)
            student_summary.append({
                "student_id": sample_row.get("student_id"),
                "student_code": sample_row.get("student_code"),
                "full_name": sample_row.get("full_name"),
                "phone": sample_row.get("phone"),
                "batch_id": sample_row.get("batch_id"),
                "batch_name": sample_row.get("batch_name"),
                "branch_id": sample_row.get("branch_id"),
                "branch_name": sample_row.get("branch_name"),
                "total_marked": 0,
                "present": 0,
                "absent": 0,
                "late": 0,
                "leave_count": 0,
                "last_marked": None,
                "expected": expected,
                "unmarked": expected,
                "leave": 0,
                "rate": 0,
                "marking_rate": 0,
            })
        student_summary.sort(key=lambda row: ((row.get("full_name") or ""), (row.get("batch_name") or "")))

        return render_template(
            "reports/attendance_monthly.html",
            report_month=report_month,
            month_label=month_label,
            start_date=start_date,
            end_date=end_date,
            calculation_end_date=calculation_end_date,
            has_calculation_window=has_calculation_window,
            branches=branches,
            selected_branch_id=selected_branch_id,
            selected_branch_name=selected_branch_name,
            can_view_all=can_view_all,
            calendar_settings=calendar_settings,
            totals=totals,
            daily_breakdown=daily_breakdown,
            branch_summary=branch_summary,
            batch_summary=batch_summary,
            student_summary=student_summary,
        )
    finally:
        conn.close()


@reports_bp.route("/attendance/unmarked-details")
@login_required
def attendance_unmarked_details():
    date_str = request.args.get("date")
    
    if not date_str:
        return redirect(url_for("reports.attendance_monthly_report"))
        
    conn = get_conn()
    cur = conn.cursor()
    try:
        branches, selected_branch_id, selected_branch_name, can_view_all = _resolve_report_branch(cur)

        calendar_settings = _get_attendance_calendar(cur, date_str, date_str)
        expected_data = _load_expected_attendance(cur, date_str, date_str, selected_branch_id, calendar_settings)
        marked_keys = _load_marked_keys(cur, date_str, date_str, selected_branch_id)
        
        # Get active trainers for dropdown
        current_inst = get_current_institute_id(default=1)
        trainers_rows = cur.execute("""
            SELECT DISTINCT u.id, u.full_name
            FROM batches b
            JOIN users u ON b.trainer_id = u.id
            JOIN branches br ON br.id = b.branch_id
            WHERE b.status = 'active' AND br.institute_id = ?
            ORDER BY u.full_name
        """, (current_inst,)).fetchall()
        trainers = [dict(r) for r in trainers_rows]

        selected_trainer_id = request.args.get("trainer_id")
        try:
            if selected_trainer_id:
                selected_trainer_id = int(selected_trainer_id)
        except ValueError:
            selected_trainer_id = None

        # Load monthly statistics of unmarked records by trainer
        selected_month = date_str[:7]
        m_report_month, m_month_label, m_start_date, m_end_date, m_last_day = _month_bounds(selected_month)
        m_calendar_settings = _get_attendance_calendar(cur, m_start_date, m_end_date)
        m_expected_data = _load_expected_attendance(cur, m_start_date, m_end_date, selected_branch_id, m_calendar_settings)
        m_marked_keys = _load_marked_keys(cur, m_start_date, m_end_date, selected_branch_id)
        
        from collections import defaultdict
        unmarked_by_trainer = defaultdict(int)
        trainer_id_map = {}
        for key in m_expected_data["keys"]:
            day, student_id, batch_id = key
            if key not in m_marked_keys:
                row = m_expected_data["row_lookup"].get((student_id, batch_id))
                if row:
                    trainer_name = row.get("trainer_name") or "Unknown Trainer"
                    trainer_id = row.get("trainer_id")
                    unmarked_by_trainer[trainer_name] += 1
                    if trainer_id:
                        trainer_id_map[trainer_name] = trainer_id

        for t in trainers:
            tname = t["full_name"]
            if tname not in unmarked_by_trainer:
                unmarked_by_trainer[tname] = 0
                trainer_id_map[tname] = t["id"]

        trainer_monthly_stats = [
            {
                "trainer_id": trainer_id_map.get(name),
                "trainer_name": name,
                "count": count
            }
            for name, count in unmarked_by_trainer.items()
        ]
        trainer_monthly_stats.sort(key=lambda x: x["count"], reverse=True)

        unmarked_students = []
        for key in expected_data["keys"]:
            day, student_id, batch_id = key
            if key not in marked_keys:
                row = expected_data["row_lookup"].get((student_id, batch_id))
                if row:
                    if selected_trainer_id and row.get("trainer_id") != selected_trainer_id:
                        continue
                    unmarked_students.append(row)
                    
        unmarked_students.sort(key=lambda r: (r.get("batch_name") or "", r.get("full_name") or ""))
        
        from datetime import datetime
        display_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%m-%Y")
        
        return render_template(
            "reports/attendance_unmarked_details.html",
            date_str=date_str,
            display_date=display_date,
            selected_branch_id=selected_branch_id,
            selected_branch_name=selected_branch_name,
            unmarked_students=unmarked_students,
            trainers=trainers,
            selected_trainer_id=selected_trainer_id,
            trainer_monthly_stats=trainer_monthly_stats,
            m_month_label=m_month_label
        )
    finally:
        conn.close()


@reports_bp.route("/attendance/quick-mark", methods=["POST"])
@login_required
def attendance_quick_mark():
    """AJAX endpoint to quickly mark attendance for a student on a specific date."""
    user_id = session.get('user_id')
    student_id = request.form.get("student_id")
    batch_id = request.form.get("batch_id")
    branch_id = request.form.get("branch_id")
    attendance_date = request.form.get("date")
    status = request.form.get("status")
    remarks = request.form.get("remarks") or ""

    if not all([student_id, batch_id, branch_id, attendance_date, status]):
        return jsonify({"success": False, "error": "Missing required fields"}), 400

    if status not in ["present", "absent", "late", "leave"]:
        return jsonify({"success": False, "error": "Invalid status"}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        current_inst = get_current_institute_id(default=1)
        cur.execute("SELECT id, branch_id, can_view_all_branches FROM users WHERE id = ?", (user_id,))
        user = cur.fetchone()
        if not user:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        # Never trust object IDs supplied by the browser. Verify that the complete
        # student/batch/branch relationship belongs to the active tenant.
        cur.execute("""
            SELECT sb.student_id
            FROM student_batches sb
            JOIN students s ON s.id = sb.student_id
            JOIN batches b ON b.id = sb.batch_id
            JOIN branches br ON br.id = b.branch_id
            WHERE sb.student_id = ? AND sb.batch_id = ? AND b.branch_id = ?
              AND s.institute_id = ? AND br.institute_id = ?
        """, (student_id, batch_id, branch_id, current_inst, current_inst))
        if not cur.fetchone():
            return jsonify({"success": False, "error": "Invalid tenant attendance target"}), 403

        can_view_all = bool(user["can_view_all_branches"]) or session.get("role") == "admin"
        if not can_view_all and int(user["branch_id"] or 0) != int(branch_id):
            return jsonify({"success": False, "error": "Branch access denied"}), 403

        # Check if record already exists
        cur.execute("""
            SELECT ar.id FROM attendance_records ar
            JOIN branches br ON br.id = ar.branch_id
            WHERE ar.batch_id = ? AND ar.student_id = ? AND ar.attendance_date = ?
              AND br.institute_id = ?
        """, (batch_id, student_id, attendance_date, current_inst))
        existing = cur.fetchone()

        now = datetime.now().isoformat(timespec="seconds")
        if existing:
            cur.execute("""
                UPDATE attendance_records
                SET status = ?, remarks = ?, marked_by = ?, updated_at = ?
                WHERE id = ?
            """, (status, remarks, user_id, now, existing["id"]))
        else:
            cur.execute("""
                INSERT INTO attendance_records (
                    attendance_date, student_id, batch_id, branch_id,
                    status, remarks, marked_by, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (attendance_date, student_id, batch_id, branch_id,
                  status, remarks, user_id, now, now))

        conn.commit()

        log_activity(user_id, branch_id, 'CREATE' if not existing else 'UPDATE',
                     'attendance', batch_id,
                     f'Quick marked attendance for student {student_id} on {attendance_date}: {status}')

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()


@reports_bp.route("/attendance/monthly/download")
@login_required
def attendance_monthly_download():
    """Download the monthly attendance report as CSV."""
    report_month, month_label, start_date, end_date, _last_day = _month_bounds(request.args.get("month"))
    _today, calculation_end_date, query_end_date, _has_calculation_window = _calculation_window(start_date, end_date)

    conn = get_conn()
    cur = conn.cursor()
    try:
        _branches, selected_branch_id, selected_branch_name, _can_view_all = _resolve_report_branch(cur)
        calendar_settings = _get_attendance_calendar(cur, start_date, end_date)
        where_clause = _attendance_where(selected_branch_id)
        params = _attendance_params(start_date, query_end_date, selected_branch_id)
        expected_data = _load_expected_attendance(cur, start_date, query_end_date, selected_branch_id, calendar_settings)
        marked_keys = _load_marked_keys(cur, start_date, query_end_date, selected_branch_id)

        cur.execute(f"""
            SELECT
                ar.attendance_date,
                br.branch_name,
                b.batch_name,
                c.course_name,
                u.full_name AS trainer_name,
                s.full_name AS student_name,
                s.student_code,
                s.phone,
                ar.status,
                ar.remarks
            FROM attendance_records ar
            JOIN students s ON ar.student_id = s.id
            LEFT JOIN batches b ON ar.batch_id = b.id
            LEFT JOIN courses c ON b.course_id = c.id
            LEFT JOIN branches br ON ar.branch_id = br.id
            LEFT JOIN users u ON b.trainer_id = u.id
            WHERE {where_clause}
            ORDER BY ar.attendance_date, br.branch_name, b.batch_name, s.full_name
        """, params)
        rows = cur.fetchall()
        unmarked_rows = []
        for attendance_date, student_id, batch_id in sorted(expected_data["keys"] - marked_keys):
            sample_row = expected_data["row_lookup"].get((student_id, batch_id), {})
            unmarked_rows.append({
                "attendance_date": attendance_date,
                "branch_name": sample_row.get("branch_name") or "",
                "batch_name": sample_row.get("batch_name") or "",
                "course_name": sample_row.get("course_name") or "",
                "trainer_name": sample_row.get("trainer_name") or "",
                "student_name": sample_row.get("full_name") or "",
                "student_code": sample_row.get("student_code") or "",
                "phone": sample_row.get("phone") or "",
                "status": "not_marked",
                "remarks": "",
            })
    finally:
        conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([f"Monthly Attendance Report - {month_label} - {selected_branch_name}"])
    writer.writerow([f"Calculated through: {calculation_end_date}"])
    writer.writerow([])
    writer.writerow(["Date", "Branch", "Batch", "Course", "Trainer", "Student", "Reg. No", "Phone", "Status", "Remarks"])
    for row in rows:
        writer.writerow([
            row["attendance_date"],
            row["branch_name"] or "",
            row["batch_name"] or "",
            row["course_name"] or "",
            row["trainer_name"] or "",
            row["student_name"] or "",
            row["student_code"] or "",
            row["phone"] or "",
            row["status"] or "",
            row["remarks"] or "",
        ])
    for row in unmarked_rows:
        writer.writerow([
            row["attendance_date"],
            row["branch_name"],
            row["batch_name"],
            row["course_name"],
            row["trainer_name"],
            row["student_name"],
            row["student_code"],
            row["phone"],
            row["status"],
            row["remarks"],
        ])
    if not rows and not unmarked_rows:
        writer.writerow(["No attendance records found for the selected month"])

    buf = io.BytesIO()
    buf.write(output.getvalue().encode("utf-8-sig"))
    buf.seek(0)
    output.close()

    branch_label = ""
    if selected_branch_id:
        branch_label = "_" + selected_branch_name.replace(" ", "_")
    filename = f"monthly_attendance_{report_month}{branch_label}.csv"
    return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)


@reports_bp.route("/attendance/settings", methods=["GET", "POST"])
@login_required
@admin_required
def attendance_calendar_settings():
    """Configure working days and attendance holidays used by monthly reports."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        _ensure_attendance_calendar_tables(cur)
        current_inst = get_current_institute_id(default=1)
        now = datetime.now(REPORT_IST).strftime("%Y-%m-%d %H:%M:%S")

        if request.method == "POST":
            action = request.form.get("action", "").strip()

            if action == "save_working_days":
                selected_days = []
                for day in request.form.getlist("working_days"):
                    if day.isdigit() and 0 <= int(day) <= 6:
                        selected_days.append(str(int(day)))
                if not selected_days:
                    flash("Please select at least one working day.", "danger")
                    return redirect(url_for("reports.attendance_calendar_settings"))

                working_days = ",".join(sorted(set(selected_days), key=int))
                cur.execute("""
                    UPDATE tenant_attendance_calendar_settings
                    SET working_days = ?, updated_at = ?
                    WHERE institute_id = ?
                """, (working_days, now, current_inst))
                conn.commit()
                flash("Attendance working days updated.", "success")
                return redirect(url_for("reports.attendance_calendar_settings"))

            if action == "add_holiday":
                holiday_date = (request.form.get("holiday_date") or "").strip()
                title = (request.form.get("title") or "").strip()
                try:
                    datetime.strptime(holiday_date, "%Y-%m-%d")
                except ValueError:
                    flash("Please enter a valid holiday date.", "danger")
                    return redirect(url_for("reports.attendance_calendar_settings"))
                if not title:
                    flash("Holiday title is required.", "danger")
                    return redirect(url_for("reports.attendance_calendar_settings"))

                cur.execute("""
                    INSERT INTO tenant_attendance_holidays
                        (institute_id, holiday_date, title, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(institute_id, holiday_date) DO UPDATE SET
                        title = excluded.title,
                        updated_at = excluded.updated_at
                """, (current_inst, holiday_date, title, now, now))
                conn.commit()
                flash("Holiday saved.", "success")
                return redirect(url_for("reports.attendance_calendar_settings"))

            if action == "delete_holiday":
                holiday_id = request.form.get("holiday_id")
                if holiday_id and holiday_id.isdigit():
                    cur.execute(
                        "DELETE FROM tenant_attendance_holidays WHERE id = ? AND institute_id = ?",
                        (int(holiday_id), current_inst),
                    )
                    conn.commit()
                    flash("Holiday deleted.", "success")
                return redirect(url_for("reports.attendance_calendar_settings"))

        settings = _get_attendance_calendar(cur)
        day_options = [
            (0, "Monday"),
            (1, "Tuesday"),
            (2, "Wednesday"),
            (3, "Thursday"),
            (4, "Friday"),
            (5, "Saturday"),
            (6, "Sunday"),
        ]
        return render_template(
            "reports/attendance_settings.html",
            settings=settings,
            day_options=day_options,
        )
    finally:
        conn.close()


@reports_bp.route("/export/<table_name>")
@login_required
@admin_required
def export_csv(table_name):
    """Export any table to CSV"""
    allowed_tables = {
        "activity_logs": "activity_logs",
        "branches": "branches",
        "courses": "courses",
        "expense_categories": "expense_categories",
        "expenses": "expenses",
        "followups": "followups",
        "installment_plans": "installment_plans",
        "invoice_items": "invoice_items",
        "invoices": "invoices",
        "leads": "leads",
        "receipts": "receipts",
        "students": "students",
        "users": "users"
    }
    
    if table_name not in allowed_tables:
        flash(f"Invalid table: {table_name}", "danger")
        return redirect(url_for("reports.dashboard"))
    
    conn = get_conn()
    cur = conn.cursor()
    
    try:
        current_inst = get_current_institute_id(default=1)
        tbl = allowed_tables[table_name]
        if tbl in ("branches", "expenses", "invoices", "leads", "receipts", "students", "installment_plans", "activity_logs", "expense_categories"):
            cur.execute(f"SELECT * FROM {tbl} WHERE institute_id = ?", (current_inst,))
        elif tbl == "invoice_items":
            cur.execute("SELECT ii.* FROM invoice_items ii JOIN invoices inv ON inv.id = ii.invoice_id WHERE inv.institute_id = ?", (current_inst,))
        elif tbl == "followups":
            cur.execute("SELECT f.* FROM followups f JOIN leads l ON l.id = f.lead_id WHERE l.institute_id = ?", (current_inst,))
        elif tbl == "users":
            cur.execute("SELECT * FROM users WHERE branch_id IN (SELECT id FROM branches WHERE institute_id = ?) OR id = ?", (current_inst, session.get("user_id")))
        else:
            cur.execute(f"SELECT * FROM {tbl}")
        rows = cur.fetchall()
        conn.close()
        
        if not rows:
            flash(f"No data in {table_name}.", "warning")
            return redirect(url_for("reports.dashboard"))
        
        columns = [description[0] for description in cur.description]
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        
        for row in rows:
            row_data = []
            for col in columns:
                value = row[col]
                if table_name == "followups" and col == "created_at" and value:
                    try:
                        if 'T' in str(value):
                            dt = datetime.fromisoformat(value)
                            value = dt.strftime("%d-%m-%Y %I:%M %p")
                    except (ValueError, AttributeError):
                        pass
                if table_name == "followups" and col == "next_followup_date" and value:
                    try:
                        dt = datetime.strptime(str(value), "%Y-%m-%d")
                        value = dt.strftime("%d-%m-%Y")
                    except (ValueError, AttributeError):
                        pass
                row_data.append(value if value is not None else "")
            writer.writerow(row_data)
        
        csv_data = output.getvalue()
        output.close()
        
        # Create response
        response_file = io.BytesIO()
        response_file.write(csv_data.encode())
        response_file.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{table_name}_{timestamp}.csv"
        
        return send_file(
            response_file,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        conn.close()
        flash(f"Error exporting {table_name}: {str(e)}", "danger")
        return redirect(url_for("reports.dashboard"))


@reports_bp.route("/export-leads-detailed")
@login_required
@admin_required
def export_leads_detailed():
    """Export detailed leads report with related data"""
    conn = get_conn()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT 
                leads.id,
                leads.name,
                leads.phone,
                leads.whatsapp,
                leads.gender,
                leads.age,
                leads.education_status,
                leads.stream,
                leads.institute_name,
                leads.career_goal,
                leads.interested_courses,
                leads.lead_source,
                leads.decision_maker,
                leads.lead_location,
                leads.start_timeframe,
                leads.lead_score,
                leads.stage,
                leads.status,
                leads.last_contact_date,
                leads.next_followup_date,
                leads.notes,
                users.full_name as assigned_to,
                leads.created_at,
                leads.updated_at
            FROM leads
            LEFT JOIN users ON leads.assigned_to_id = users.id
            WHERE leads.is_deleted = 0
            ORDER BY leads.created_at DESC
        """)
        
        rows = cur.fetchall()
        conn.close()
        
        if not rows:
            flash("No leads data to export.", "warning")
            return redirect(url_for("reports.dashboard"))
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        headers = [
            "ID", "Name", "Phone", "WhatsApp", "Gender", "Age", 
            "Education Status", "Stream", "Institute", "Career Goal",
            "Interested Courses", "Lead Source", "Decision Maker", "Lead Location",
            "Start Timeframe", "Lead Score", "Stage", "Status",
            "Last Contact", "Next Follow-up", "Notes", "Assigned To", "Created", "Updated"
        ]
        writer.writerow(headers)
        
        for row in rows:
            writer.writerow([
                row["id"],
                row["name"],
                row["phone"],
                row["whatsapp"] or "",
                row["gender"] or "",
                row["age"] or "",
                row["education_status"] or "",
                row["stream"] or "",
                row["institute_name"] or "",
                row["career_goal"] or "",
                row["interested_courses"] or "",
                row["lead_source"] or "",
                row["decision_maker"] or "",
                row["lead_location"] or "",
                row["start_timeframe"] or "",
                row["lead_score"] or "",
                row["stage"],
                row["status"],
                row["last_contact_date"] or "",
                row["next_followup_date"] or "",
                row["notes"] or "",
                row["assigned_to"] or "",
                row["created_at"],
                row["updated_at"] or ""
            ])
        
        csv_data = output.getvalue()
        output.close()
        
        response_file = io.BytesIO()
        response_file.write(csv_data.encode())
        response_file.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"leads_detailed_{timestamp}.csv"
        
        return send_file(
            response_file,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        conn.close()
        flash(f"Error exporting leads: {str(e)}", "danger")
        return redirect(url_for("reports.dashboard"))


@reports_bp.route("/export-students-detailed")
@login_required
@admin_required
def export_students_detailed():
    """Export detailed students report"""
    conn = get_conn()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT 
                students.id,
                students.student_code,
                students.full_name,
                students.phone,
                students.email,
                students.gender,
                students.address,
                students.education_level,
                students.qualification,
                students.student_location,
                students.employment_status,
                students.status,
                branches.branch_name,
                students.joined_date,
                students.created_at
            FROM students
            LEFT JOIN branches ON students.branch_id = branches.id
            ORDER BY students.created_at DESC
        """)
        
        rows = cur.fetchall()
        conn.close()
        
        if not rows:
            flash("No students data to export.", "warning")
            return redirect(url_for("reports.dashboard"))
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        headers = [
            "ID", "Student Code", "Full Name", "Phone", "Email", "Gender",
            "Address", "Education Level", "Qualification", "Student Location",
            "Employment Status", "Status", "Branch", "Joined Date", "Created"
        ]
        writer.writerow(headers)
        
        for row in rows:
            writer.writerow([
                row["id"],
                row["student_code"],
                row["full_name"],
                row["phone"],
                row["email"] or "",
                row["gender"] or "",
                row["address"] or "",
                row["education_level"] or "",
                row["qualification"] or "",
                row["student_location"] or "",
                row["employment_status"] or "",
                row["status"],
                row["branch_name"] or "",
                row["joined_date"],
                row["created_at"]
            ])
        
        csv_data = output.getvalue()
        output.close()
        
        response_file = io.BytesIO()
        response_file.write(csv_data.encode())
        response_file.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"students_detailed_{timestamp}.csv"
        
        return send_file(
            response_file,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        conn.close()
        flash(f"Error exporting students: {str(e)}", "danger")
        return redirect(url_for("reports.dashboard"))


@reports_bp.route("/import")
@login_required
@admin_required
def import_page():
    """CSV Import Management Page"""
    return render_template("reports/import.html")


@reports_bp.route("/sample/<table_name>")
@login_required
@admin_required
def download_sample(table_name):
    """Generate sample CSV or Excel file for data import with automatic Reference Sheets."""
    file_type = request.args.get("file_type", "xlsx").strip().lower()
    
    samples = {
        "branches": {
            "headers": ["branch_name", "branch_code", "address", "is_active"],
            "rows": [
                ["Head Office", "HO", "Main branch address", "1"],
                ["Branch 1", "B1", "Branch 1 address", "1"],
            ]
        },
        "courses": {
            "headers": ["course_name", "duration", "fee", "course_type", "is_active"],
            "rows": [
                ["Tally", "45 Days", "5000", "standard", "1"],
                ["Excel Advanced", "30 Days", "4000", "standard", "1"],
            ]
        },
        "leads": {
            "headers": ["name", "phone", "whatsapp", "gender", "age", "education_status", "stream", "institute_name", "career_goal", "interested_courses", "lead_source", "decision_maker", "lead_location", "start_timeframe", "lead_score", "stage", "status", "lost_reason", "last_contact_date", "next_followup_date", "followup_count", "notes", "assigned_to_id"],
            "rows": [
                ["John Doe", "9876543210", "9876543210", "Male", "25", "Graduate", "Commerce", "ABC Institute", "Job", "Tally,Excel", "Walk-in", "Self", "urban", "Immediately", "8", "New Lead", "active", "", "21-03-2026", "28-03-2026", "1", "Interested in Tally", "1"],
                ["Jane Smith", "9123456789", "9123456789", "Female", "22", "School", "Science", "XYZ School", "Skill Development", "Excel,Power BI", "Referral", "Parents", "rural", "Within 1 Month", "7", "Converted", "active", "", "20-03-2026", "27-03-2026", "3", "Converted to student", ""],
            ]
        },
        "students": {
            "headers": [
                "branch_name", "student_code", "full_name", "phone", "email", "gender", "date_of_birth",
                "father_name", "mother_name", "parent_name", "parent_contact",
                "address", "locality", "city", "state", "pincode", "landmark", "alternate_phone", "address_type",
                "education_level", "qualification",
                "tenth_institution", "tenth_board", "tenth_year", "tenth_percentage",
                "puc_institution", "puc_board", "puc_stream", "puc_year", "puc_percentage",
                "degree_institution", "degree_university", "degree_course", "degree_year", "degree_percentage",
                "student_location", "employment_status", "status", "joined_date"
            ],
            "rows": [
                [
                    "Main Branch", "1515001", "Student Name", "9876543210", "student@example.com", "Male", "2002-05-15",
                    "Father Name", "Mother Name", "Parent Name", "9876543211",
                    "MG Road, 4th Cross", "Indiranagar", "Bengaluru", "Karnataka", "560038", "Near Metro Station", "9876543212", "Home",
                    "Undergraduate", "BE",
                    "St. Joseph High School", "SSLC", "2018", "85%",
                    "National PU College", "PU Board", "Science", "2020", "88%",
                    "PES University", "VTU", "B.E Computer Science", "2024", "8.5 CGPA",
                    "urban", "student", "active", "21-03-2026"
                ],
                [
                    "Main Branch", "1515002", "Another Student", "9123456789", "student2@example.com", "Female", "2004-08-20",
                    "Father Name 2", "Mother Name 2", "Parent Name 2", "9123456780",
                    "Station Road", "Town Area", "Mysore", "Karnataka", "570001", "Opposite Bus Stand", "9123456781", "Home",
                    "Pre-University", "12th",
                    "Government School", "State Board", "2020", "78%",
                    "Government PU College", "PU Board", "Commerce", "2022", "80%",
                    "", "", "", "", "",
                    "rural", "unemployed", "active", "21-03-2026"
                ],
            ]
        },
        "invoices": {
            "headers": ["invoice_number", "student_id", "invoice_date", "subtotal", "discount_type", "discount_value", "discount_amount", "total_amount", "installment_type", "notes", "status", "created_by", "branch_id"],
            "rows": [
                ["GIT/B/001", "1", "21-03-2026", "5000", "percentage", "10", "500", "4500", "full", "Course Fee", "unpaid", "1", "1"],
                ["GIT/B/002", "2", "20-03-2026", "4000", "fixed", "300", "300", "3700", "installment", "Excel training", "unpaid", "1", "1"],
            ]
        },
        "receipts": {
            "headers": ["receipt_number", "invoice_id", "receipt_date", "amount_received", "payment_mode", "notes"],
            "rows": [
                ["GIT/RCP/001", "1", "21-03-2026", "5000", "cash", "Full payment"],
                ["GIT/RCP/002", "2", "20-03-2026", "2000", "bank_transfer", "First installment"],
            ]
        },
        "installments": {
            "headers": ["invoice_id", "installment_number", "due_date", "amount", "status"],
            "rows": [
                ["1", "1", "21-04-2026", "2500", "pending"],
                ["1", "2", "21-05-2026", "2500", "pending"],
            ]
        },
        "expenses": {
            "headers": ["expense_type", "category", "amount", "description", "expense_date", "branch_id"],
            "rows": [
                ["rent", "office", "20000", "Monthly office rent", "21-03-2026", "1"],
                ["utilities", "office", "5000", "Electricity bill", "21-03-2026", "1"],
            ]
        },
        "activity_logs": {
            "headers": ["user_id", "branch_id", "action_type", "module_name", "record_id", "description"],
            "rows": [
                ["1", "1", "create", "leads", "1", "Created new lead"],
                ["1", "1", "update", "students", "1", "Updated student record"],
            ]
        },
        "expense_categories": {
            "headers": ["category_name", "is_active"],
            "rows": [
                ["Rent", "1"],
                ["Utilities", "1"],
                ["Office Supplies", "1"],
            ]
        },
        "followups": {
            "headers": ["lead_id", "user_id", "method", "outcome", "note", "next_followup_date", "created_at"],
            "rows": [
                ["1", "", "call", "interested", "Discussed Tally course, interested in classes", "28-03-2026", "23-03-2026 02:30 PM"],
                ["1", "", "whatsapp", "callback_later", "Sent course details, waiting for response", "31-03-2026", "22-03-2026 10:15 AM"],
                ["2", "", "email", "not_interested", "Student declined, pursuing other options", "", "21-03-2026 04:45 PM"],
                ["3", "", "walk_in", "converted", "Student enrolled in Excel course", "", "20-03-2026 09:00 AM"],
            ]
        },
        "installment_plans": {
            "headers": ["invoice_id", "installment_no", "due_date", "amount_due", "amount_paid", "status", "remarks"],
            "rows": [
                ["1", "1", "21-04-2026", "2500", "2500", "paid", "First payment received"],
                ["1", "2", "21-05-2026", "2500", "0", "pending", ""],
            ]
        },
        "invoice_items": {
            "headers": ["invoice_id", "course_id", "description", "quantity", "unit_price", "discount", "line_total"],
            "rows": [
                ["1", "1", "Tally Course", "1", "5000", "0", "5000"],
                ["2", "2", "Excel Advanced", "1", "4000", "200", "3800"],
            ]
        },
        "users": {
            "headers": ["full_name", "username", "role", "phone", "branch_id", "can_view_all_branches", "is_active"],
            "rows": [
                ["Admin User", "admin", "admin", "9876543210", "1", "1", "1"],
                ["Staff User", "staff", "staff", "9123456789", "1", "0", "1"],
            ]
        },
    }
    
    if table_name not in samples:
        flash(f"No sample available for {table_name}", "warning")
        return redirect(url_for("reports.import_page"))
    
    sample = samples[table_name]
    conn = get_conn()
    current_inst = get_current_institute_id(default=1)

    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = f"{table_name.capitalize()} Data"
        
        # Style headers
        ws_data.append(sample["headers"])
        header_fill = openpyxl.styles.PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for cell in ws_data[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in sample["rows"]:
            ws_data.append(row)
            
        # Add Reference Sheet for Branches
        branches = conn.execute("SELECT id, branch_name, branch_code, address FROM branches WHERE institute_id = ? ORDER BY id", (current_inst,)).fetchall()
        if branches:
            ws_b = wb.create_sheet(title="Branches Reference")
            ws_b.append(["Branch ID", "Branch Name", "Branch Code", "Address"])
            ref_fill = openpyxl.styles.PatternFill(start_color="198754", end_color="198754", fill_type="solid")
            for cell in ws_b[1]:
                cell.fill = ref_fill
                cell.font = header_font
            for b in branches:
                ws_b.append([b["id"], b["branch_name"], b["branch_code"], b["address"]])

        # Add Reference Sheet for Courses
        courses = conn.execute("SELECT id, course_name, duration, fee FROM courses WHERE institute_id = ? ORDER BY id", (current_inst,)).fetchall()
        if courses:
            ws_c = wb.create_sheet(title="Courses Reference")
            ws_c.append(["Course ID", "Course Name", "Duration", "Fee"])
            course_fill = openpyxl.styles.PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
            for cell in ws_c[1]:
                cell.fill = course_fill
                cell.font = header_font
            for c in courses:
                ws_c.append([c["id"], c["course_name"], c["duration"], c["fee"]])

        conn.close()

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{table_name}_sample.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    else:
        conn.close()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(sample["headers"])
        for row in sample["rows"]:
            writer.writerow(row)
        
        csv_data = output.getvalue()
        output.close()
        
        response_file = io.BytesIO()
        response_file.write(csv_data.encode("utf-8"))
        response_file.seek(0)
        
        filename = f"{table_name}_sample.csv"
        return send_file(
            response_file,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename
        )


@reports_bp.route("/upload", methods=["POST"])
@login_required
@admin_required
def upload_csv():
    """Handle Excel (.xlsx/.xls) or CSV file upload and import"""
    
    file = request.files.get("csv_file") or request.files.get("file")
    if not file or not file.filename:
        flash("No file selected", "danger")
        return redirect(url_for("reports.import_page"))
    
    table_name = request.form.get("table_name", "").strip()
    if not table_name:
        flash("No table selected", "danger")
        return redirect(url_for("reports.import_page"))
    
    allowed_tables = ["activity_logs", "branches", "courses", "expense_categories", "expenses", "followups", "installment_plans", "invoice_items", "invoices", "leads", "receipts", "students", "users"]
    if table_name not in allowed_tables:
        flash(f"Invalid table: {table_name}", "danger")
        return redirect(url_for("reports.import_page"))
    
    try:
        file_content = file.read()
        if not file_content:
            flash("❌ Uploaded file is empty. Please select a valid Excel or CSV file.", "danger")
            return redirect(url_for("reports.import_page"))
        
        raw_rows = []
        filename_lower = file.filename.lower()

        if filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheet = wb.active  # Uses the primary data sheet
                rows_iter = list(sheet.iter_rows(values_only=True))
                if not rows_iter or len(rows_iter) < 2:
                    flash("❌ Excel file has no data rows. First row must be headers.", "danger")
                    return redirect(url_for("reports.import_page"))
                
                raw_headers = [str(h).replace('\ufeff', '').strip() if h is not None else '' for h in rows_iter[0]]
                for row_vals in rows_iter[1:]:
                    if not any(v is not None and str(v).strip() for v in row_vals):
                        continue
                    row_dict = {}
                    for idx, h in enumerate(raw_headers):
                        if h:
                            val = row_vals[idx] if idx < len(row_vals) else None
                            row_dict[h] = str(val).strip() if val is not None else ""
                    raw_rows.append(row_dict)
            except Exception as e:
                flash(f"❌ Unable to parse Excel file: {str(e)}", "danger")
                return redirect(url_for("reports.import_page"))
        else:
            try:
                text_content = file_content.decode('utf-8')
            except UnicodeDecodeError:
                try:
                    text_content = file_content.decode('latin-1')
                except UnicodeDecodeError:
                    flash("❌ File encoding error. Please save your file in UTF-8 format.", "danger")
                    return redirect(url_for("reports.import_page"))
            
            stream = io.StringIO(text_content)
            reader = csv.DictReader(stream)
            if not reader.fieldnames:
                flash("❌ CSV file has no headers. First row must contain column names.", "danger")
                return redirect(url_for("reports.import_page"))
            
            fieldnames = [name.replace('\ufeff', '').strip() if name else name for name in reader.fieldnames]
            for row in reader:
                normalized_row = {k.replace('\ufeff', '').strip() if k else k: (v.strip() if v else "") for k, v in row.items()}
                if any(normalized_row.values()):
                    raw_rows.append(normalized_row)

        if not raw_rows:
            flash("❌ No valid data rows found in uploaded file.", "danger")
            return redirect(url_for("reports.import_page"))

        conn = get_conn()
        cur = conn.cursor()
        current_inst = get_current_institute_id(default=1)
        
        rows_imported = 0
        errors = []
        
        for idx, row in enumerate(raw_rows, start=2):
            if not row or not any(row.values()):
                continue
            
            try:
                if table_name == "branches":
                    is_active = int(row.get("is_active", 1))
                    if is_active:
                        lock_and_check_limit(conn, current_inst, "branches")
                    cur.execute("""
                        INSERT INTO branches (
                            institute_id, branch_name, branch_code, address, is_active, created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        row.get("branch_name", "").strip(),
                        row.get("branch_code", "").strip(),
                        row.get("address", "").strip(),
                        is_active,
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
                
                elif table_name == "courses":
                    cur.execute("""
                        INSERT INTO courses (
                            institute_id, course_name, duration, fee, course_type, is_active, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        row.get("course_name", "").strip(),
                        row.get("duration", "").strip(),
                        float(row.get("fee", 0)) if row.get("fee") else 0,
                        row.get("course_type", "standard").strip(),
                        int(row.get("is_active", 1)),
                        datetime.now().isoformat(timespec="seconds"),
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
                
                elif table_name == "leads":
                    # Validate lead_location field
                    lead_location = row.get("lead_location", "").strip() if row.get("lead_location") else None
                    if lead_location and lead_location.lower() not in ['rural', 'urban']:
                        error_msg = f"Row {idx + 1}: lead_location must be 'rural' or 'urban' (got '{lead_location}')"
                        errors.append(error_msg)
                        continue
                    
                    # Normalize location to lowercase
                    if lead_location:
                        lead_location = lead_location.lower()
                    
                    # Handle assigned_to_id (optional, defaults to current user)
                    assigned_to_id = None
                    if row.get("assigned_to_id"):
                        try:
                            assigned_to_id = int(row.get("assigned_to_id"))
                        except (ValueError, TypeError):
                            error_msg = f"Row {idx + 1}: assigned_to_id must be a valid user ID (got '{row.get('assigned_to_id')}')"
                            errors.append(error_msg)
                            continue
                    else:
                        assigned_to_id = session.get("user_id")
                    
                    cur.execute("""
                        INSERT INTO leads (
                            institute_id, name, phone, whatsapp, gender, age, education_status, stream,
                            institute_name, career_goal, interested_courses, lead_source, decision_maker, 
                            lead_location, start_timeframe, lead_score, stage, status, lost_reason,
                            last_contact_date, next_followup_date, followup_count, notes, 
                            is_deleted, assigned_to_id, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        row.get("name", "").strip(),
                        row.get("phone", "").strip(),
                        row.get("whatsapp", "").strip() or None,
                        row.get("gender", "").strip() or None,
                        int(row.get("age", 0)) if row.get("age") else None,
                        row.get("education_status", "").strip() or None,
                        row.get("stream", "").strip() or None,
                        row.get("institute_name", "").strip() or None,
                        row.get("career_goal", "").strip() or None,
                        row.get("interested_courses", "").strip() or None,
                        row.get("lead_source", "").strip() or None,
                        row.get("decision_maker", "Self").strip() or "Self",
                        lead_location,
                        row.get("start_timeframe", "").strip() or None,
                        int(row.get("lead_score", 0)) if row.get("lead_score") else None,
                        row.get("stage", "New Lead").strip() or "New Lead",
                        row.get("status", "active").strip() or "active",
                        row.get("lost_reason", "").strip() or None,
                        parse_date(row.get("last_contact_date", "")) or None,
                        parse_date(row.get("next_followup_date", "")) or None,
                        int(row.get("followup_count", 0)) if row.get("followup_count") else 0,
                        row.get("notes", "").strip() or None,
                        0,
                        assigned_to_id,
                        datetime.now().isoformat(timespec="seconds"),
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
                
                elif table_name == "students":
                    student_location = row.get("student_location", "").strip() if row.get("student_location") else None
                    if student_location and student_location.lower() not in ['rural', 'urban']:
                        error_msg = f"Row {idx}: student_location must be 'rural' or 'urban' (got '{student_location}')"
                        errors.append(error_msg)
                        continue
                    if student_location:
                        student_location = student_location.lower()
                    
                    student_status = row.get("status", "active").strip() or "active"
                    if student_status == "active":
                        lock_and_check_limit(conn, current_inst, "students")

                    # Smart Branch Resolution (branch_id, branch_name, or branch_code)
                    branch_id = None
                    raw_b_id = str(row.get("branch_id", "")).strip()
                    raw_b_name = str(row.get("branch_name", "")).strip()
                    raw_b_code = str(row.get("branch_code", "")).strip()

                    if raw_b_id and raw_b_id.isdigit():
                        b_chk = conn.execute("SELECT id FROM branches WHERE id = ? AND institute_id = ?", (int(raw_b_id), current_inst)).fetchone()
                        if b_chk:
                            branch_id = b_chk["id"]

                    if not branch_id and raw_b_name:
                        b_chk = conn.execute("SELECT id FROM branches WHERE (LOWER(branch_name) = LOWER(?) OR LOWER(branch_code) = LOWER(?)) AND institute_id = ?", (raw_b_name, raw_b_name, current_inst)).fetchone()
                        if b_chk:
                            branch_id = b_chk["id"]

                    if not branch_id and raw_b_code:
                        b_chk = conn.execute("SELECT id FROM branches WHERE LOWER(branch_code) = LOWER(?) AND institute_id = ?", (raw_b_code, current_inst)).fetchone()
                        if b_chk:
                            branch_id = b_chk["id"]

                    if not branch_id:
                        branch_id = session.get("branch_id")
                        if branch_id:
                            b_chk = conn.execute("SELECT id FROM branches WHERE id = ? AND institute_id = ?", (branch_id, current_inst)).fetchone()
                            if not b_chk:
                                branch_id = None
                        if not branch_id:
                            b_first = conn.execute("SELECT id FROM branches WHERE institute_id = ? ORDER BY id ASC LIMIT 1", (current_inst,)).fetchone()
                            if b_first:
                                branch_id = b_first["id"]

                    if not branch_id:
                        errors.append(f"Row {idx}: Unable to determine valid branch for this institute")
                        continue

                    full_name = row.get("full_name", "").strip()
                    phone = row.get("phone", "").strip()
                    if not full_name:
                        errors.append(f"Row {idx}: full_name is required")
                        continue
                    if not phone:
                        errors.append(f"Row {idx}: phone is required")
                        continue

                    student_code = row.get("student_code", "").strip()
                    if not student_code:
                        max_id_row = conn.execute("SELECT MAX(id) AS max_id FROM students").fetchone()
                        next_seq = (max_id_row["max_id"] or 0) + 1
                        student_code = f"S{current_inst}{next_seq:04d}"

                    cur.execute("""
                        INSERT INTO students (
                            institute_id, student_code, full_name, phone, email, gender, date_of_birth,
                            father_name, mother_name, parent_name, parent_contact,
                            address, locality, city, state, pincode, landmark, alternate_phone, address_type,
                            education_level, qualification,
                            tenth_institution, tenth_board, tenth_year, tenth_percentage,
                            puc_institution, puc_board, puc_stream, puc_year, puc_percentage,
                            degree_institution, degree_university, degree_course, degree_year, degree_percentage,
                            student_location, employment_status, status, branch_id, joined_date, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        student_code,
                        full_name,
                        phone,
                        row.get("email", "").strip() or None,
                        row.get("gender", "").strip() or None,
                        parse_date(row.get("date_of_birth", "")) or None,
                        row.get("father_name", "").strip() or None,
                        row.get("mother_name", "").strip() or None,
                        row.get("parent_name", "").strip() or None,
                        row.get("parent_contact", "").strip() or None,
                        row.get("address", "").strip() or None,
                        row.get("locality", "").strip() or None,
                        row.get("city", "").strip() or None,
                        row.get("state", "").strip() or None,
                        row.get("pincode", "").strip() or None,
                        row.get("landmark", "").strip() or None,
                        row.get("alternate_phone", "").strip() or None,
                        row.get("address_type", "Home").strip() or "Home",
                        row.get("education_level", "").strip() or None,
                        row.get("qualification", "").strip() or None,
                        row.get("tenth_institution", "").strip() or None,
                        row.get("tenth_board", "").strip() or None,
                        row.get("tenth_year", "").strip() or None,
                        row.get("tenth_percentage", "").strip() or None,
                        row.get("puc_institution", "").strip() or None,
                        row.get("puc_board", "").strip() or None,
                        row.get("puc_stream", "").strip() or None,
                        row.get("puc_year", "").strip() or None,
                        row.get("puc_percentage", "").strip() or None,
                        row.get("degree_institution", "").strip() or None,
                        row.get("degree_university", "").strip() or None,
                        row.get("degree_course", "").strip() or None,
                        row.get("degree_year", "").strip() or None,
                        row.get("degree_percentage", "").strip() or None,
                        student_location,
                        row.get("employment_status", "unemployed").strip() or "unemployed",
                        student_status,
                        branch_id,
                        parse_date(row.get("joined_date", "")) or datetime.now().isoformat(timespec="seconds"),
                        datetime.now().isoformat(timespec="seconds"),
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
                
                elif table_name == "invoices":
                    try:
                        # Validate required fields for 13-column invoice import
                        invoice_number = row.get("invoice_number", "").strip()
                        student_id = row.get("student_id", "").strip()
                        invoice_date = row.get("invoice_date", "").strip()
                        subtotal = row.get("subtotal", "0").strip()
                        discount_type = row.get("discount_type", "none").strip().lower()
                        discount_value = row.get("discount_value", "0").strip()
                        discount_amount = row.get("discount_amount", "0").strip()
                        total_amount = row.get("total_amount", "0").strip()
                        installment_type = row.get("installment_type", "full").strip().lower()
                        notes = row.get("notes", "").strip()
                        status = row.get("status", "unpaid").strip().lower()
                        created_by = row.get("created_by", "").strip()
                        branch_id = row.get("branch_id", "").strip()
                        
                        # Validation
                        if not invoice_number:
                            errors.append(f"Row {idx}: invoice_number is required")
                            continue
                        if not student_id:
                            errors.append(f"Row {idx}: student_id is required")
                            continue
                        if not invoice_date:
                            errors.append(f"Row {idx}: invoice_date is required")
                            continue
                        
                        # Validate and parse date (supports DD-MM-YYYY or YYYY-MM-DD)
                        parsed_invoice_date = parse_date(invoice_date)
                        if not parsed_invoice_date:
                            errors.append(f"Row {idx}: invalid invoice_date format (use DD-MM-YYYY or YYYY-MM-DD)")
                            continue
                        invoice_date = parsed_invoice_date
                        
                        # Validate numbers
                        try:
                            subtotal = float(subtotal)
                            discount_value = float(discount_value)
                            discount_amount = float(discount_amount)
                            total_amount = float(total_amount)
                        except ValueError as e:
                            errors.append(f"Row {idx}: invalid number format - {str(e)}")
                            continue
                        
                        # Convert IDs and validate tenant scoping
                        try:
                            student_id = int(student_id)
                            created_by = int(created_by) if created_by else session.get("user_id")
                            branch_id = int(branch_id)
                        except ValueError:
                            errors.append(f"Row {idx}: student_id, created_by, and branch_id must be valid integers")
                            continue

                        # Verify student_id belongs to current institute
                        if not conn.execute(
                            "SELECT id FROM students WHERE id = ? AND institute_id = ?",
                            (student_id, current_inst),
                        ).fetchone():
                            errors.append(f"Row {idx}: student_id {student_id} does not exist in this institute")
                            continue

                        # Verify branch_id belongs to current institute
                        if not conn.execute(
                            "SELECT id FROM branches WHERE id = ? AND institute_id = ?",
                            (branch_id, current_inst),
                        ).fetchone():
                            errors.append(f"Row {idx}: branch_id {branch_id} does not exist in this institute")
                            continue
                        
                        # Insert invoice with institute_id
                        cur.execute("""
                            INSERT INTO invoices (
                                institute_id, invoice_no, student_id, invoice_date, subtotal, 
                                discount_type, discount_value, discount_amount, total_amount,
                                installment_type, notes, status, created_by, branch_id, created_at
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            current_inst, invoice_number, student_id, invoice_date, subtotal,
                            discount_type, discount_value, discount_amount, total_amount,
                            installment_type, notes, status, created_by, branch_id,
                            datetime.now().isoformat(timespec="seconds")
                        ))
                        rows_imported += 1
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                        continue
                
                elif table_name == "receipts":
                    receipt_invoice_id = int(row.get("invoice_id", 0)) if row.get("invoice_id") else 0
                    receipt_amount = float(row.get("amount_received", 0)) if row.get("amount_received") else 0
                    
                    if receipt_invoice_id <= 0 or not conn.execute(
                        "SELECT id FROM invoices WHERE id = ? AND institute_id = ?",
                        (receipt_invoice_id, current_inst),
                    ).fetchone():
                        errors.append(f"Row {idx}: invoice_id {receipt_invoice_id} does not exist in this institute")
                        continue

                    cur.execute("""
                        INSERT INTO receipts (
                            institute_id, receipt_no, invoice_id, receipt_date, amount_received, payment_mode, notes, created_by, created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        row.get("receipt_number", "").strip(),
                        receipt_invoice_id,
                        parse_date(row.get("receipt_date", "")) or None,
                        receipt_amount,
                        row.get("payment_mode", "cash").strip() or "cash",
                        row.get("notes", "").strip() or None,
                        session.get("user_id"),
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    
                    # Update invoice status based on total receipts
                    cur.execute("""
                        SELECT total_amount FROM invoices WHERE id = ? AND institute_id = ?
                    """, (receipt_invoice_id, current_inst))
                    invoice_row = cur.fetchone()
                    
                    if invoice_row:
                        invoice_total = float(invoice_row["total_amount"] or 0)
                        
                        cur.execute("""
                            SELECT IFNULL(SUM(amount_received), 0) AS total_received
                            FROM receipts
                            WHERE invoice_id = ? AND institute_id = ?
                        """, (receipt_invoice_id, current_inst))
                        total_received = float(cur.fetchone()["total_received"] or 0)
                        
                        # Determine new status
                        if total_received >= invoice_total:
                            new_status = "paid"
                        elif total_received > 0:
                            new_status = "partially_paid"
                        else:
                            new_status = "unpaid"
                        
                        cur.execute("""
                            UPDATE invoices
                            SET status = ?, updated_at = ?
                            WHERE id = ? AND institute_id = ?
                        """, (new_status, datetime.now().isoformat(timespec="seconds"), receipt_invoice_id, current_inst))
                    
                    rows_imported += 1
                
                elif table_name == "installments" or table_name == "installment_plans":
                    inv_id = int(row.get("invoice_id", 0)) if row.get("invoice_id") else 0
                    if inv_id <= 0 or not conn.execute(
                        "SELECT id FROM invoices WHERE id = ? AND institute_id = ?",
                        (inv_id, current_inst),
                    ).fetchone():
                        errors.append(f"Row {idx}: invoice_id {inv_id} does not exist in this institute")
                        continue

                    inst_no = int(row.get("installment_number") or row.get("installment_no") or 0)
                    amt_due = float(row.get("amount") or row.get("amount_due") or 0)
                    amt_paid = float(row.get("amount_paid", 0)) if row.get("amount_paid") else 0
                    
                    cur.execute("""
                        INSERT INTO installment_plans (
                            invoice_id, installment_no, due_date, amount_due, amount_paid, status, remarks, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        inv_id,
                        inst_no,
                        parse_date(row.get("due_date", "")) or None,
                        amt_due,
                        amt_paid,
                        row.get("status", "pending").strip() or "pending",
                        row.get("remarks", "").strip() or None,
                        datetime.now().isoformat(timespec="seconds"),
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
                
                elif table_name == "activity_logs":
                    log_branch_id = int(row.get("branch_id", session.get("branch_id") or 1)) if row.get("branch_id") else session.get("branch_id")
                    cur.execute("""
                        INSERT INTO activity_logs (
                            institute_id, user_id, branch_id, action_type, module_name, record_id, description, created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        int(row.get("user_id", session.get("user_id"))) if row.get("user_id") else session.get("user_id"),
                        log_branch_id,
                        row.get("action_type", "").strip() or "import",
                        row.get("module_name", "").strip() or "import_export",
                        int(row.get("record_id", 0)) if row.get("record_id") else None,
                        row.get("description", "").strip(),
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
                
                elif table_name == "expense_categories":
                    cur.execute("""
                        INSERT INTO expense_categories (institute_id, category_name, is_active, created_at)
                        VALUES (?, ?, ?, ?)
                    """, (
                        current_inst,
                        row.get("category_name", "").strip(),
                        int(row.get("is_active", 1)),
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
                
                elif table_name == "followups":
                    # Skip completely empty rows
                    if not any(row.values()):
                        continue
                    
                    # Validate lead_id exists
                    lead_id_str = row.get("lead_id", "").strip()
                    if not lead_id_str:
                        available_cols = ", ".join([k for k in row.keys() if k])
                        non_empty_values = {k: v for k, v in row.items() if v and v.strip()}
                        error_msg = f"Row {idx}: lead_id is required (available: {available_cols} | data: {non_empty_values})"
                        errors.append(error_msg)
                        continue
                    
                    try:
                        lead_id = int(lead_id_str)
                    except ValueError:
                        error_msg = f"Row {idx + 1}: lead_id must be a valid number (got '{lead_id_str}')"
                        errors.append(error_msg)
                        continue
                    
                    # Check if lead exists in current institute
                    cur.execute("SELECT id FROM leads WHERE id = ? AND institute_id = ? AND is_deleted = 0", (lead_id, current_inst))
                    if not cur.fetchone():
                        error_msg = f"Row {idx + 1}: Lead ID {lead_id} not found in this institute"
                        errors.append(error_msg)
                        continue
                    
                    # Validate user_id if provided
                    user_id = session.get("user_id")
                    if row.get("user_id", "").strip():
                        try:
                            user_id = int(row.get("user_id"))
                            cur.execute("SELECT id FROM users WHERE id = ? AND institute_id = ? AND is_active = 1", (user_id, current_inst))
                            if not cur.fetchone():
                                error_msg = f"Row {idx + 1}: User ID {user_id} not found in this institute"
                                errors.append(error_msg)
                                continue
                        except ValueError:
                            error_msg = f"Row {idx + 1}: user_id must be a valid number (got '{row.get('user_id')}')"
                            errors.append(error_msg)
                            continue
                    
                    # Insert followup
                    created_at_str = row.get("created_at", "").strip()
                    if created_at_str:
                        try:
                            created_at_parsed = datetime.strptime(created_at_str, "%d-%m-%Y %I:%M %p")
                            created_at_value = created_at_parsed.isoformat()
                        except ValueError:
                            try:
                                created_at_parsed = datetime.strptime(created_at_str, "%d-%m-%Y %H:%M")
                                created_at_value = created_at_parsed.isoformat()
                            except ValueError:
                                try:
                                    created_at_parsed = datetime.strptime(created_at_str, "%d-%b-%Y %H:%M")
                                    created_at_value = created_at_parsed.isoformat()
                                except ValueError:
                                    try:
                                        created_at_parsed = datetime.strptime(created_at_str, "%d-%b-%Y %I:%M %p")
                                        created_at_value = created_at_parsed.isoformat()
                                    except ValueError:
                                        try:
                                            created_at_parsed = datetime.strptime(created_at_str, "%d-%m-%Y")
                                            created_at_value = created_at_parsed.isoformat()
                                        except ValueError:
                                            errors.append(f"Row {rows_imported + 1}: Invalid created_at format '{created_at_str}', using current timestamp")
                                            created_at_value = datetime.now().isoformat(timespec="seconds")
                    else:
                        created_at_value = datetime.now().isoformat(timespec="seconds")
                    
                    cur.execute("""
                        INSERT INTO followups (
                            institute_id, lead_id, user_id, method, outcome, note, next_followup_date, created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        lead_id,
                        user_id,
                        row.get("method", "").strip() or None,
                        row.get("outcome", "").strip() or None,
                        row.get("note", "").strip() or None,
                        parse_date(row.get("next_followup_date", "")) if row.get("next_followup_date", "").strip() else None,
                        created_at_value
                    ))
                    rows_imported += 1
                
                elif table_name == "invoice_items":
                    inv_id = int(row.get("invoice_id", 0)) if row.get("invoice_id") else 0
                    if inv_id <= 0 or not conn.execute(
                        "SELECT id FROM invoices WHERE id = ? AND institute_id = ?",
                        (inv_id, current_inst),
                    ).fetchone():
                        errors.append(f"Row {idx}: invoice_id {inv_id} does not exist in this institute")
                        continue

                    course_id = int(row.get("course_id")) if row.get("course_id") else None
                    if course_id and not conn.execute(
                        "SELECT id FROM courses WHERE id = ? AND institute_id = ?",
                        (course_id, current_inst),
                    ).fetchone():
                        errors.append(f"Row {idx}: course_id {course_id} does not exist in this institute")
                        continue

                    cur.execute("""
                        INSERT INTO invoice_items (
                            invoice_id, course_id, description, quantity, unit_price, discount, line_total, created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        inv_id,
                        course_id,
                        row.get("description", "").strip(),
                        int(row.get("quantity", 1)) if row.get("quantity") else 1,
                        float(row.get("unit_price", 0)) if row.get("unit_price") else 0,
                        float(row.get("discount", 0)) if row.get("discount") else 0,
                        float(row.get("line_total", 0)) if row.get("line_total") else 0,
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
                
                elif table_name == "users":
                    user_active = int(row.get("is_active", 1))
                    if user_active:
                        lock_and_check_limit(conn, current_inst, "staff")
                    branch_id = (
                        int(row.get("branch_id"))
                        if row.get("branch_id") else session.get("branch_id")
                    )
                    if branch_id and not conn.execute(
                        "SELECT id FROM branches WHERE id = ? AND institute_id = ?",
                        (branch_id, current_inst),
                    ).fetchone():
                        errors.append(f"Row {idx}: branch_id must belong to this institute")
                        continue
                    cur.execute("""
                        INSERT INTO users (
                            institute_id, full_name, username, password_hash, role, phone, branch_id,
                            can_view_all_branches, is_active, created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        row.get("full_name", "").strip(),
                        row.get("username", "").strip(),
                        generate_password_hash(row.get("username", "")),  # Default password = username
                        row.get("role", "staff").strip() or "staff",
                        row.get("phone", "").strip() or None,
                        branch_id,
                        int(row.get("can_view_all_branches", 0)),
                        user_active,
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    imported_user_id = cur.lastrowid
                    cur.execute(
                        """INSERT INTO institute_memberships (
                               institute_id, user_id, membership_role, is_active,
                               created_at, updated_at
                           ) VALUES (?, ?, ?, ?, ?, ?)""",
                        (
                            current_inst, imported_user_id,
                            "institute_admin"
                            if (row.get("role", "staff").strip() or "staff") == "admin"
                            else "staff",
                            user_active,
                            datetime.now().isoformat(timespec="seconds"),
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )
                    rows_imported += 1
                
                elif table_name == "expenses":
                    exp_branch_id = int(row.get("branch_id", session.get("branch_id") or 1)) if row.get("branch_id") else session.get("branch_id")
                    if exp_branch_id and not conn.execute(
                        "SELECT id FROM branches WHERE id = ? AND institute_id = ?",
                        (exp_branch_id, current_inst),
                    ).fetchone():
                        errors.append(f"Row {idx}: branch_id {exp_branch_id} does not exist in this institute")
                        continue

                    cur.execute("""
                        INSERT INTO expenses (
                            institute_id, expense_type, category, amount, description, expense_date, branch_id, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        current_inst,
                        row.get("expense_type", "").strip() or "other",
                        row.get("category", "").strip() or None,
                        float(row.get("amount", 0)) if row.get("amount") else 0,
                        row.get("description", "").strip() or None,
                        parse_date(row.get("expense_date", "")) or datetime.now().isoformat(timespec="seconds"),
                        exp_branch_id,
                        datetime.now().isoformat(timespec="seconds"),
                        datetime.now().isoformat(timespec="seconds")
                    ))
                    rows_imported += 1
            
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        conn.commit()
        
        log_activity(
            user_id=session.get("user_id"),
            branch_id=session.get("branch_id"),
            action_type="import",
            module_name="reports",
            record_id=None,
            description=f"Imported {rows_imported} rows to {table_name}"
        )
        
        conn.close()
        
        if errors:
            error_message = f"✓ Imported {rows_imported} rows, but {len(errors)} row(s) failed:\n"
            for err in errors[:3]:
                error_message += f"\n  • {err}"
            if len(errors) > 3:
                error_message += f"\n  ... and {len(errors) - 3} more errors"
            flash(error_message, "warning")
        else:
            flash(f"✅ Successfully imported {rows_imported} rows into {table_name}!", "success")
        
        return redirect(url_for("reports.import_page"))
    
    except Exception as e:
        error_detail = str(e)
        
        # Provide helpful error messages
        if "no column named" in error_detail.lower():
            flash(f"❌ Column mismatch error.\n\nThe CSV file has a column that doesn't exist in the {table_name} table.\n\nError: {error_detail}\n\n📋 Tip: Download the sample CSV to see correct column names.", "danger")
        elif "constraint" in error_detail.lower() or "foreign key" in error_detail.lower():
            flash(f"❌ Data relationship error.\n\nA record references data that doesn't exist (e.g., student_id without student).\n\nError: {error_detail}", "danger")
        elif "not null" in error_detail.lower():
            flash(f"❌ Missing required data.\n\nA required field is empty or missing.\n\nError: {error_detail}\n\n📋 Tip: Check that all required columns have values.", "danger")
        else:
            flash(f"❌ Error importing CSV:\n\n{error_detail}\n\n📋 Try downloading a sample CSV and comparing your format.", "danger")
        
        return redirect(url_for("reports.import_page"))


@reports_bp.route("/lms-attendance-gap", methods=["GET"])
@login_required
def lms_attendance_gap():
    """Report showing students who attend physical classes but are lagging in LMS progress."""
    user_id = session.get('user_id')
    conn = get_conn()
    try:
        cur = conn.cursor()
        current_inst = get_current_institute_id(default=1)
        
        # Determine accessible branch context
        cur.execute("SELECT id, branch_id, can_view_all_branches, role FROM users WHERE id = ?", (user_id,))
        user = cur.fetchone()
        if not user:
            flash("User session not found", "danger")
            return redirect(url_for("core.dashboard"))

        # Extract filters
        f_branch_id = request.args.get("branch_id", "").strip()
        f_batch_id = request.args.get("batch_id", "").strip()
        f_program_id = request.args.get("program_id", "").strip()
        f_min_gap = request.args.get("min_gap", "3").strip()
        
        selected_branch_id = int(f_branch_id) if f_branch_id.isdigit() else None
        selected_batch_id = int(f_batch_id) if f_batch_id.isdigit() else None
        selected_program_id = int(f_program_id) if f_program_id.isdigit() else None
        min_gap = int(f_min_gap) if f_min_gap.isdigit() else 3

        # Scoping/Permissions check
        role = user["role"]
        user_branch_id = user["branch_id"]
        can_view_all = user["can_view_all_branches"] or (role == 'admin')

        if not can_view_all:
            # Force user's branch
            selected_branch_id = user_branch_id

        # Load dropdown filters
        if can_view_all:
            branches = cur.execute(
                "SELECT id, branch_name FROM branches WHERE is_active = 1 AND institute_id = ? ORDER BY branch_name",
                (current_inst,)
            ).fetchall()
        else:
            branches = cur.execute(
                "SELECT id, branch_name FROM branches WHERE id = ? AND is_active = 1 AND institute_id = ?",
                (user_branch_id, current_inst)
            ).fetchall()

        if selected_branch_id:
            batches = cur.execute(
                "SELECT id, batch_name FROM batches WHERE branch_id = ? AND status = 'active' ORDER BY batch_name",
                (selected_branch_id,)
            ).fetchall()
        elif not can_view_all:
            batches = cur.execute(
                "SELECT id, batch_name FROM batches WHERE branch_id = ? AND status = 'active' ORDER BY batch_name",
                (user_branch_id,)
            ).fetchall()
        else:
            batches = cur.execute(
                "SELECT id, batch_name FROM batches WHERE status = 'active' AND branch_id IN (SELECT id FROM branches WHERE institute_id = ?) ORDER BY batch_name",
                (current_inst,)
            ).fetchall()

        programs = cur.execute(
            "SELECT id, program_name FROM lms_programs WHERE is_active = 1 AND is_deleted = 0 AND institute_id = ? ORDER BY program_name",
            (current_inst,)
        ).fetchall()

        # Enforce program enrollment explicit check
        _enroll_check = """(
            EXISTS (
                SELECT 1 FROM lms_student_program_access spa
                WHERE spa.student_id = s.id AND spa.program_id = lp.id
                  AND spa.is_active = 1
                  AND (spa.access_end_date IS NULL OR spa.access_end_date >= date('now'))
            )
        )"""

        where_clauses = [
            "s.status = 'active'",
            "s.institute_id = ?",
            "lp.is_active = 1",
            "lp.is_deleted = 0",
            "lp.institute_id = ?",
            _enroll_check,
            "EXISTS (SELECT 1 FROM student_batches sb_act WHERE sb_act.student_id = s.id AND sb_act.status = 'active')"
        ]
        params = [current_inst, current_inst]

        if selected_program_id:
            where_clauses.append("lp.id = ?")
            params.append(selected_program_id)

        if selected_batch_id:
            where_clauses.append("""EXISTS (
                SELECT 1 FROM student_batches sb_f
                WHERE sb_f.student_id = s.id AND sb_f.batch_id = ? AND sb_f.status = 'active'
            )""")
            params.append(selected_batch_id)

        if selected_branch_id:
            where_clauses.append("""(
                EXISTS (
                    SELECT 1 FROM student_batches sb_br
                    JOIN batches bat_br ON bat_br.id = sb_br.batch_id
                    WHERE sb_br.student_id = s.id AND sb_br.status = 'active'
                      AND bat_br.branch_id = ?
                )
                OR s.branch_id = ?
            )""")
            params.extend([selected_branch_id, selected_branch_id])

        where_sql = " AND ".join(where_clauses)

        _master_check = """EXISTS (
            SELECT 1 FROM lms_program_chapters pcx
            JOIN lms_master_chapters mcx ON mcx.id = pcx.master_chapter_id
            JOIN lms_master_topics   mtx ON mtx.master_chapter_id = mcx.id
            WHERE pcx.program_id = lp.id AND pcx.is_visible = 1
              AND mcx.status = 'active' AND mtx.status = 'active'
        )"""

        # Query all student/program pairings
        sql = f"""
            SELECT
                s.id               AS student_id,
                s.student_code,
                s.full_name,
                lp.id              AS program_id,
                lp.program_name,
                COALESCE(b.batch_name,  '') AS batch_name,
                COALESCE(br.branch_name, br2.branch_name, '') AS branch_name,
                COALESCE(u.full_name,   '') AS trainer_name,
                -- 1. Classes attended (unique days marked present/late)
                (
                    SELECT COUNT(DISTINCT ar.attendance_date)
                    FROM attendance_records ar
                    WHERE ar.student_id = s.id AND ar.status IN ('present', 'late')
                ) AS classes_attended,
                -- 2. Total topics
                CASE WHEN {_master_check} THEN (
                    SELECT COUNT(*)
                    FROM lms_master_topics mt
                    JOIN lms_program_chapters pc ON pc.master_chapter_id = mt.master_chapter_id
                    JOIN lms_master_chapters  mc ON mc.id = pc.master_chapter_id
                    WHERE pc.program_id = lp.id AND pc.is_visible = 1
                      AND mc.status = 'active'  AND mt.status = 'active'
                ) ELSE (
                    SELECT COUNT(*)
                    FROM lms_topics lt
                    JOIN lms_chapters lc ON lt.chapter_id = lc.id
                    WHERE lc.program_id = lp.id AND lt.is_active = 1
                ) END AS total_topics,
                -- 3. Completed topics
                CASE WHEN {_master_check} THEN (
                    SELECT COUNT(*)
                    FROM lms_master_topic_progress mtp
                    JOIN lms_master_topics mt ON mt.id = mtp.master_topic_id
                    JOIN lms_program_chapters pc ON pc.master_chapter_id = mt.master_chapter_id
                    JOIN lms_master_chapters  mc ON mc.id = pc.master_chapter_id
                    WHERE mtp.student_id = s.id AND mtp.program_id = lp.id AND mtp.is_completed = 1
                      AND pc.program_id = lp.id AND pc.is_visible = 1
                      AND mc.status = 'active'  AND mt.status = 'active'
                ) ELSE (
                    SELECT COUNT(*)
                    FROM lms_topic_progress tp
                    JOIN lms_topics lt ON tp.topic_id = lt.id
                    JOIN lms_chapters lc ON lt.chapter_id = lc.id
                    WHERE tp.student_id = s.id AND lc.program_id = lp.id AND tp.is_completed = 1
                ) END AS completed_topics,
                -- last activity
                (
                    SELECT MAX(last_act) FROM (
                        SELECT MAX(tp.completed_at) AS last_act
                        FROM lms_topic_progress tp
                        JOIN lms_topics lt ON tp.topic_id = lt.id
                        JOIN lms_chapters lc ON lt.chapter_id = lc.id
                        WHERE tp.student_id = s.id AND lc.program_id = lp.id
                        UNION ALL
                        SELECT MAX(mtp.completed_at) AS last_act
                        FROM lms_master_topic_progress mtp
                        WHERE mtp.student_id = s.id AND mtp.program_id = lp.id
                    )
                ) AS last_activity
            FROM students s
            JOIN lms_programs lp ON lp.is_active = 1 AND lp.is_deleted = 0
            LEFT JOIN batches b ON b.id = COALESCE(
                (SELECT spa2.batch_id FROM lms_student_program_access spa2
                 WHERE spa2.student_id = s.id AND spa2.program_id = lp.id
                   AND spa2.is_active = 1 AND spa2.batch_id IS NOT NULL LIMIT 1),
                (SELECT MIN(sb3.batch_id) FROM student_batches sb3
                 WHERE sb3.student_id = s.id AND sb3.status = 'active')
            )
            LEFT JOIN branches br  ON br.id  = b.branch_id
            LEFT JOIN branches br2 ON br2.id = s.branch_id
            LEFT JOIN users    u   ON u.id   = b.trainer_id
            WHERE {where_sql}
        """

        raw_rows = cur.execute(sql, params).fetchall()
        
        # Filter and compute stats
        defaulters = []
        for r in raw_rows:
            attended = r["classes_attended"] or 0
            completed = r["completed_topics"] or 0
            total = r["total_topics"] or 0
            
            # Only consider students who have actually attended some classes
            if attended == 0:
                continue
                
            gap = attended - completed
            if gap >= min_gap:
                progress_pct = round((completed / total * 100), 1) if total > 0 else 0.0
                
                # Format last activity date
                last_act_str = "—"
                if r["last_activity"]:
                    try:
                        from datetime import datetime as _dt
                        last_act_str = _dt.fromisoformat(str(r["last_activity"]).replace('T', ' ').split('.')[0]).strftime('%d %b %Y %I:%M %p')
                    except Exception:
                        last_act_str = str(r["last_activity"])[:16]

                defaulters.append({
                    "student_id": r["student_id"],
                    "student_code": r["student_code"],
                    "full_name": r["full_name"],
                    "batch_name": r["batch_name"],
                    "branch_name": r["branch_name"],
                    "trainer_name": r["trainer_name"],
                    "program_id": r["program_id"],
                    "program_name": r["program_name"],
                    "classes_attended": attended,
                    "completed_topics": completed,
                    "total_topics": total,
                    "gap": gap,
                    "progress_pct": progress_pct,
                    "last_activity": last_act_str
                })

        # Sort by gap descending (highest gap first)
        defaulters.sort(key=lambda x: x["gap"], reverse=True)

        from datetime import datetime
        m_month_label = datetime.now().strftime("%B %Y")

        return render_template(
            "reports/lms_attendance_gap.html",
            defaulters=defaulters,
            branches=branches,
            batches=batches,
            programs=programs,
            selected_branch_id=selected_branch_id,
            selected_batch_id=selected_batch_id,
            selected_program_id=selected_program_id,
            min_gap=min_gap,
            can_view_all=can_view_all,
            m_month_label=m_month_label
        )
    finally:
        conn.close()
