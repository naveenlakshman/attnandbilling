from flask import Blueprint, render_template, session, flash, send_file, request, redirect, url_for
from functools import wraps
from io import BytesIO, StringIO
from datetime import datetime
import sqlite3
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from db import get_conn, log_activity
from config import DB_PATH
from modules.core.utils import login_required, admin_required

import_export_bp = Blueprint("import_export", __name__)


from services.tenant_context import get_current_institute_id


def get_all_tables_data(current_inst=1):
    """
    Retrieves all tables and their data from the database.
    Returns a dictionary where keys are table names and values are lists of dictionaries.
    """
    conn = get_conn()
    cursor = conn.cursor()
    
    # Check if MySQL or SQLite
    import pymysql
    is_mysql = hasattr(conn, '_conn') and isinstance(conn._conn, pymysql.connections.Connection)
    
    if is_mysql:
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()
        table_names = [list(row.values())[0] for row in tables]
    else:
        cursor.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name NOT LIKE 'sqlite_%'
            ORDER BY name
        """)
        table_names = [row[0] for row in cursor.fetchall()]
    
    tables_data = {}
    
    for table_name in table_names:
        # Get column names
        if is_mysql:
            cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
            cols_info = cursor.fetchall()
            columns = [col['Field'] for col in cols_info]
        else:
            cursor.execute(f"PRAGMA table_info([{table_name}])")
            cols_info = cursor.fetchall()
            columns = [col[1] for col in cols_info]

        where_clause = ""
        params = []
        if table_name == "institutes":
            where_clause = "WHERE id = ?"
            params = [current_inst]
        elif "institute_id" in columns:
            where_clause = "WHERE institute_id = ?"
            params = [current_inst]
        elif table_name in ("installment_plans", "invoice_items", "invoice_payments"):
            where_clause = "WHERE invoice_id IN (SELECT id FROM invoices WHERE institute_id = ?)"
            params = [current_inst]
        elif table_name in ("leave_requests", "student_batches", "student_notes", "student_documents", "student_qualifications", "lms_student_topic_progress", "lms_topic_progress"):
            where_clause = "WHERE student_id IN (SELECT id FROM students WHERE institute_id = ?)"
            params = [current_inst]
        elif table_name == "followups":
            where_clause = "WHERE lead_id IN (SELECT id FROM leads WHERE institute_id = ?)"
            params = [current_inst]
        elif table_name == "attendance_records":
            where_clause = "WHERE student_id IN (SELECT id FROM students WHERE institute_id = ?) OR batch_id IN (SELECT id FROM batches WHERE branch_id IN (SELECT id FROM branches WHERE institute_id = ?))"
            params = [current_inst, current_inst]
        elif table_name in ("lms_program_chapters", "lms_chapters"):
            where_clause = "WHERE program_id IN (SELECT id FROM lms_programs WHERE institute_id = ?)"
            params = [current_inst]
        elif table_name == "lms_topics":
            where_clause = "WHERE chapter_id IN (SELECT lc.id FROM lms_chapters lc WHERE lc.program_id IN (SELECT id FROM lms_programs WHERE institute_id = ?))"
            params = [current_inst]
        elif "branch_id" in columns:
            where_clause = "WHERE branch_id IN (SELECT id FROM branches WHERE institute_id = ?)"
            params = [current_inst]

        cursor.execute(f"SELECT * FROM `{table_name}` {where_clause}", params)
        rows = cursor.fetchall()
        
        # Convert rows to list of dictionaries
        data = []
        if rows:
            for row in rows:
                row_dict = dict(row)
                for k, v in row_dict.items():
                    if isinstance(v, (bytes, bytearray)):
                        row_dict[k] = "<binary>"
                    elif hasattr(v, 'isoformat'):
                        row_dict[k] = str(v)
                data.append(row_dict)

        # Fallback for company_profile when empty for tenant
        if table_name == "company_profile" and not data:
            from db import get_company_profile
            prof = get_company_profile(current_inst)
            if prof:
                prof_dict = dict(prof)
                for k, v in prof_dict.items():
                    if isinstance(v, (bytes, bytearray)):
                        prof_dict[k] = "<binary>"
                    elif hasattr(v, 'isoformat'):
                        prof_dict[k] = str(v)
                data.append(prof_dict)
                for k in prof_dict.keys():
                    if k not in columns:
                        columns.append(k)
        
        tables_data[table_name] = {
            'columns': columns,
            'data': data
        }
    
    conn.close()
    return tables_data



def create_excel_workbook(tables_data):
    """
    Creates an Excel workbook with multiple sheets, one for each table.
    Each sheet contains the table data with headers.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet
    
    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for table_name, table_info in tables_data.items():
        # Create a new sheet for each table
        sheet = workbook.create_sheet(title=table_name[:31])  # Excel sheet name limit is 31 chars
        
        # Add headers
        columns = table_info['columns']
        for col_idx, column_name in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data rows
        for row_idx, row_data in enumerate(table_info['data'], 2):
            for col_idx, column_name in enumerate(columns, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                val = row_data.get(column_name)
                if isinstance(val, (bytes, bytearray)):
                    val = "<binary>"
                elif hasattr(val, 'isoformat'):
                    val = str(val)
                cell.value = val
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        
        # Adjust column widths
        for col_idx, column_name in enumerate(columns, 1):
            max_length = len(str(column_name))
            for row_data in table_info['data']:
                cell_value = str(row_data.get(column_name, ''))
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
            sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = adjusted_width
    
    return workbook


@import_export_bp.route("/")
@login_required
@admin_required
def import_export_dashboard():
    """Import/Export dashboard page"""
    return render_template("import_export/dashboard.html")


@import_export_bp.route("/export/all-tables", methods=["GET"])
@login_required
@admin_required
def export_all_tables():
    """
    Export all database tables to a single Excel workbook.
    Each table is in a separate sheet with the table name as sheet name.
    """
    try:
        current_inst = get_current_institute_id(default=1)
        # Get all tables and their data
        tables_data = get_all_tables_data(current_inst=current_inst)
        
        if not tables_data:
            flash("No tables found in the database.", "warning")
            return redirect(url_for("import_export.import_export_dashboard"))
        
        # Create Excel workbook
        workbook = create_excel_workbook(tables_data)
        
        # Save to bytes
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Database_Export_{timestamp}.xlsx"
        
        # Send file to user
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f"Error exporting data: {str(e)}", "danger")
        return redirect(url_for("import_export.import_export_dashboard"))

